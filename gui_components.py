"""
gui_components.py
Modular GUI components for Analog Garage Workbench.
Enhanced with comprehensive scrollbar support at all levels.

Version 3.1
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
from typing import Callable, Optional, Dict, Any, List
import os


# Import configuration
from config import (
    ADI_COLORS, UI_CONFIG, FORM_OPTIONS, FORM_DEFAULTS, LOGO_PATH,
    OUTPUT_TYPES, OUTPUT_TYPE_DISPLAY_TO_ID
)


# =============================================================================
# ENHANCED SCROLLABLE FRAME COMPONENT
# =============================================================================

class ScrollableFrame(tk.Frame):
    """
    A scrollable frame container with both vertical and optional horizontal scrolling.
    Scrollbars automatically appear/disappear based on content size.
    """
    
    def __init__(
        self, 
        parent, 
        bg_color: str = None, 
        show_horizontal: bool = False,
        always_show_vertical: bool = False,
        **kwargs
    ):
        super().__init__(parent, **kwargs)
        
        self.bg_color = bg_color or ADI_COLORS["white"]
        self.show_horizontal = show_horizontal
        self.always_show_vertical = always_show_vertical
        self.configure(bg=self.bg_color)
        
        # Create main container
        self.container = tk.Frame(self, bg=self.bg_color)
        self.container.pack(fill=tk.BOTH, expand=True)
        
        # Create canvas
        self.canvas = tk.Canvas(
            self.container,
            bg=self.bg_color,
            highlightthickness=0,
            borderwidth=0
        )
        
        # Create vertical scrollbar
        self.v_scrollbar = ttk.Scrollbar(
            self.container,
            orient="vertical",
            command=self.canvas.yview
        )
        
        # Create horizontal scrollbar (if enabled)
        if self.show_horizontal:
            self.h_scrollbar = ttk.Scrollbar(
                self.container,
                orient="horizontal",
                command=self.canvas.xview
            )
        
        # Create inner frame
        self.inner_frame = tk.Frame(self.canvas, bg=self.bg_color)
        
        # Create window in canvas
        self.canvas_window = self.canvas.create_window(
            (0, 0),
            window=self.inner_frame,
            anchor="nw"
        )
        
        # Configure canvas scrolling
        self.canvas.configure(yscrollcommand=self._on_v_scroll)
        if self.show_horizontal:
            self.canvas.configure(xscrollcommand=self._on_h_scroll)
        
        # Layout
        if self.show_horizontal:
            self.h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Show vertical scrollbar based on setting
        self.v_scrollbar_visible = False
        if self.always_show_vertical:
            self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.v_scrollbar_visible = True
        
        # Bind events
        self.inner_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        
        # Bind mouse wheel
        self._bind_mousewheel()
        
        # H scrollbar visibility tracking
        self.h_scrollbar_visible = self.show_horizontal
    
    def _on_v_scroll(self, *args):
        """Handle vertical scrollbar movement."""
        self.v_scrollbar.set(*args)
        if not self.always_show_vertical:
            self._update_v_scrollbar_visibility()
    
    def _on_h_scroll(self, *args):
        """Handle horizontal scrollbar movement."""
        if self.show_horizontal:
            self.h_scrollbar.set(*args)
    
    def _on_frame_configure(self, event=None):
        """Update scroll region when inner frame size changes."""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        if not self.always_show_vertical:
            self._update_v_scrollbar_visibility()
    
    def _on_canvas_configure(self, event):
        """Update inner frame width when canvas is resized."""
        if not self.show_horizontal:
            # Make inner frame fill canvas width
            self.canvas.itemconfig(self.canvas_window, width=event.width)
        if not self.always_show_vertical:
            self._update_v_scrollbar_visibility()
    
    def _update_v_scrollbar_visibility(self):
        """Show/hide vertical scrollbar based on content size."""
        bbox = self.canvas.bbox("all")
        if bbox:
            content_height = bbox[3] - bbox[1]
            canvas_height = self.canvas.winfo_height()
            
            if content_height > canvas_height and canvas_height > 1:
                if not self.v_scrollbar_visible:
                    self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                    self.v_scrollbar_visible = True
            else:
                if self.v_scrollbar_visible and not self.always_show_vertical:
                    self.v_scrollbar.pack_forget()
                    self.v_scrollbar_visible = False
    
    def _bind_mousewheel(self):
        """Bind mouse wheel events for scrolling."""
        self.inner_frame.bind("<Enter>", self._on_enter)
        self.inner_frame.bind("<Leave>", self._on_leave)
    
    def _on_enter(self, event):
        """Enable mouse wheel scrolling when mouse enters."""
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)
        if self.show_horizontal:
            self.canvas.bind_all("<Shift-MouseWheel>", self._on_h_mousewheel)
    
    def _on_leave(self, event):
        """Disable mouse wheel scrolling when mouse leaves."""
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")
        if self.show_horizontal:
            self.canvas.unbind_all("<Shift-MouseWheel>")
    
    def _on_mousewheel(self, event):
        """Handle vertical mouse wheel scrolling."""
        if self.v_scrollbar_visible:
            if event.delta:
                delta = -1 * (event.delta // 120)
            else:
                delta = -1 * event.delta
            self.canvas.yview_scroll(int(delta), "units")
    
    def _on_h_mousewheel(self, event):
        """Handle horizontal mouse wheel scrolling (Shift+Wheel)."""
        if self.show_horizontal and self.h_scrollbar_visible:
            if event.delta:
                delta = -1 * (event.delta // 120)
            else:
                delta = -1 * event.delta
            self.canvas.xview_scroll(int(delta), "units")
    
    def _on_mousewheel_linux(self, event):
        """Handle mouse wheel scrolling (Linux)."""
        if self.v_scrollbar_visible:
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")
    
    def scroll_to_top(self):
        """Scroll to the top."""
        self.canvas.yview_moveto(0)
    
    def scroll_to_bottom(self):
        """Scroll to the bottom."""
        self.canvas.yview_moveto(1)
    
    def get_frame(self) -> tk.Frame:
        """Return the inner frame where content should be added."""
        return self.inner_frame


# =============================================================================
# COLLAPSIBLE SECTION COMPONENT
# =============================================================================

class CollapsibleSection(tk.Frame):
    """
    A collapsible section with header and content area.
    Can be expanded/collapsed by clicking the header.
    """
    
    def __init__(
        self,
        parent,
        title: str,
        bg_color: str = None,
        header_bg: str = None,
        initially_expanded: bool = True,
        scrollable: bool = True,
        scroll_height: int = 200,
        **kwargs
    ):
        super().__init__(parent, **kwargs)
        
        self.bg_color = bg_color or ADI_COLORS["white"]
        self.header_bg = header_bg or ADI_COLORS["light_blue"]
        self.title = title
        self.is_expanded = initially_expanded
        self.scrollable = scrollable
        self.scroll_height = scroll_height
        
        self.configure(bg=self.bg_color)
        self._create_widgets()
    
    def _create_widgets(self):
        # Header
        self.header = tk.Frame(self, bg=self.header_bg, cursor="hand2")
        self.header.pack(fill=tk.X)
        
        self.toggle_label = tk.Label(
            self.header,
            text=f"{'▼' if self.is_expanded else '▶'} {self.title}",
            font=(UI_CONFIG["font_family"], 11, "bold"),
            bg=self.header_bg,
            fg=ADI_COLORS["dark_blue"],
            padx=10,
            pady=8
        )
        self.toggle_label.pack(side=tk.LEFT)
        
        # Bind click events
        self.header.bind("<Button-1>", self._toggle)
        self.toggle_label.bind("<Button-1>", self._toggle)
        
        # Content container
        self.content_container = tk.Frame(self, bg=self.bg_color)
        
        if self.scrollable:
            # Scrollable content
            self.scroll_frame = ScrollableFrame(
                self.content_container,
                bg_color=self.bg_color,
                always_show_vertical=True
            )
            self.scroll_frame.pack(fill=tk.BOTH, expand=True)
            self.content = self.scroll_frame.get_frame()
            
            # Set a maximum height for the scroll area
            self.content_container.configure(height=self.scroll_height)
        else:
            self.content = tk.Frame(self.content_container, bg=self.bg_color)
            self.content.pack(fill=tk.BOTH, expand=True)
        
        if self.is_expanded:
            self.content_container.pack(fill=tk.BOTH, expand=True)
    
    def _toggle(self, event=None):
        """Toggle section expansion."""
        if self.is_expanded:
            self.content_container.pack_forget()
            self.toggle_label.config(text=f"▶ {self.title}")
            self.is_expanded = False
        else:
            self.content_container.pack(fill=tk.BOTH, expand=True)
            self.toggle_label.config(text=f"▼ {self.title}")
            self.is_expanded = True
    
    def get_content_frame(self) -> tk.Frame:
        """Return the content frame for adding widgets."""
        return self.content
    
    def expand(self):
        """Expand the section."""
        if not self.is_expanded:
            self._toggle()
    
    def collapse(self):
        """Collapse the section."""
        if self.is_expanded:
            self._toggle()


# =============================================================================
# HEADER COMPONENT
# =============================================================================

class HeaderComponent(tk.Frame):
    """Application header with logo and title."""
    
    def __init__(self, parent, demo_mode: bool = True):
        super().__init__(parent, bg=ADI_COLORS["primary_blue"])
        self.logo_photo = None
        self._create_widgets(demo_mode)
    
    def _create_widgets(self, demo_mode: bool):
        self.config(pady=6)  # CHANGED: Reduced from 12 to 6
        
        logo_loaded = self._load_logo()
        
        if not logo_loaded:
            self._create_text_logo()
        
        title = tk.Label(
            self,
            text="Analog Garage",
            font=(UI_CONFIG["font_family"], 18, "bold"),  # CHANGED: Reduced from 24 to 18
            bg=ADI_COLORS["primary_blue"],
            fg="white"
        )
        title.pack(pady=(5, 0))  # CHANGED: Reduced from (10, 0)
        
        subtitle = tk.Label(
            self,
            text="Prompt Generator + Script Executor",
            font=(UI_CONFIG["font_family"], 11),  # CHANGED: Reduced from 13 to 11
            bg=ADI_COLORS["primary_blue"],
            fg=ADI_COLORS["light_blue"]
        )
        subtitle.pack()
        
        if demo_mode:
            self._create_demo_badge()

    
    def _load_logo(self) -> bool:
        try:
            from PIL import Image, ImageTk
            
            if os.path.exists(LOGO_PATH):
                logo_image = Image.open(LOGO_PATH)
                logo_image = logo_image.resize((200, 60), Image.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_image)
                
                logo_label = tk.Label(self, image=self.logo_photo, bg=ADI_COLORS["primary_blue"])
                logo_label.pack(pady=(0, 8))
                return True
        except ImportError:
            pass
        except Exception:
            pass
        return False
    
    def _create_text_logo(self):
        container = tk.Frame(self, bg=ADI_COLORS["primary_blue"])
        container.pack()
        
        tk.Label(
            container,
            text="◆ ANALOG",
            font=(UI_CONFIG["font_family"], UI_CONFIG["title_font_size"], "bold"),
            bg=ADI_COLORS["primary_blue"],
            fg="white"
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Label(
            container,
            text="DEVICES",
            font=(UI_CONFIG["font_family"], UI_CONFIG["title_font_size"]),
            bg=ADI_COLORS["primary_blue"],
            fg=ADI_COLORS["light_blue"]
        ).pack(side=tk.LEFT)
    
    def _create_demo_badge(self):
        container = tk.Frame(self, bg=ADI_COLORS["primary_blue"], pady=8)
        container.pack()
        
        tk.Label(
            container,
            text="⚡ DEMO VERSION",
            font=(UI_CONFIG["font_family"], 9, "bold"),
            bg=ADI_COLORS["accent_orange"],
            fg="white",
            padx=12,
            pady=3
        ).pack()


# =============================================================================
# STATUS BAR COMPONENT
# =============================================================================

class StatusBarComponent(tk.Frame):
    """Application status bar."""
    
    def __init__(self, parent, version: str = "1.0"):
        super().__init__(parent, bg=ADI_COLORS["dark_blue"])
        self.version = version
        self._create_widgets()
    
    def _create_widgets(self):
        self.config(pady=8)
        
        self.status_label = tk.Label(
            self,
            text="Ready • Use Tab 1 to generate prompts, Tab 2 to execute scripts",
            bg=ADI_COLORS["dark_blue"],
            fg=ADI_COLORS["light_blue"],
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"])
        )
        self.status_label.pack(side=tk.LEFT, padx=15)
        
        version_label = tk.Label(
            self,
            text=f"v{self.version} | Analog Garage",
            bg=ADI_COLORS["dark_blue"],
            fg=ADI_COLORS["text_gray"],
            font=(UI_CONFIG["font_family"], 9)
        )
        version_label.pack(side=tk.RIGHT, padx=15)
    
    def update_status(self, message: str, color: Optional[str] = None):
        self.status_label.config(text=message)
        if color:
            self.status_label.config(fg=color)


# =============================================================================
# OUTPUT TYPE SELECTOR COMPONENT
# =============================================================================

class OutputTypeSelectorComponent(tk.Frame):
    """Component for selecting the output type (Excel/PowerPoint/Word)."""
    
    def __init__(self, parent, on_change: Optional[Callable] = None):
        super().__init__(parent, bg=ADI_COLORS["white"])
        self.on_change = on_change
        self._create_widgets()
    
    def _create_widgets(self):
        # Container with visual emphasis
        container = tk.Frame(self, bg=ADI_COLORS["light_blue"], padx=15, pady=12)
        container.pack(fill=tk.X, padx=10, pady=10)
        
        # Title row
        title_frame = tk.Frame(container, bg=ADI_COLORS["light_blue"])
        title_frame.pack(fill=tk.X)
        
        tk.Label(
            title_frame,
            text="🎯 Select Output Type",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["light_blue"],
            fg=ADI_COLORS["dark_blue"]
        ).pack(side=tk.LEFT)
        
        # Description
        tk.Label(
            container,
            text="Choose the type of deliverable you want to generate:",
            font=(UI_CONFIG["font_family"], 10),
            bg=ADI_COLORS["light_blue"],
            fg=ADI_COLORS["dark_gray"]
        ).pack(anchor="w", pady=(5, 10))
        
        # Dropdown
        dropdown_frame = tk.Frame(container, bg=ADI_COLORS["light_blue"])
        dropdown_frame.pack(fill=tk.X)
        
        output_options = FORM_OPTIONS.get("output_type", ["Excel Value Creation Model"])
        
        self.output_type_combo = ttk.Combobox(
            dropdown_frame,
            width=45,
            state="readonly",
            values=output_options,
            font=(UI_CONFIG["font_family"], 11)
        )
        
        if output_options:
            self.output_type_combo.set(output_options[0])
        
        self.output_type_combo.pack(side=tk.LEFT)
        
        # Bind change event
        self.output_type_combo.bind("<<ComboboxSelected>>", self._on_selection_change)
        
        # Description label
        self.desc_label = tk.Label(
            container,
            text=self._get_description(),
            font=(UI_CONFIG["font_family"], 9, "italic"),
            bg=ADI_COLORS["light_blue"],
            fg=ADI_COLORS["text_gray"],
            wraplength=500,
            justify=tk.LEFT
        )
        self.desc_label.pack(anchor="w", pady=(8, 0))
    
    def _on_selection_change(self, event=None):
        """Handle selection change."""
        self.desc_label.config(text=self._get_description())
        if self.on_change:
            self.on_change(self.get_output_type_id())
    
    def _get_description(self) -> str:
        """Get description for current selection."""
        output_id = self.get_output_type_id()
        if output_id and output_id in OUTPUT_TYPES:
            return OUTPUT_TYPES[output_id]["description"]
        return ""
    
    def get_output_type_id(self) -> str:
        """Get the selected output type ID."""
        display_name = self.output_type_combo.get()
        return OUTPUT_TYPE_DISPLAY_TO_ID.get(display_name, "excel_value_model")
    
    def get_output_type_config(self) -> dict:
        """Get full configuration for selected output type."""
        output_id = self.get_output_type_id()
        return OUTPUT_TYPES.get(output_id, OUTPUT_TYPES.get("excel_value_model", {}))
    
    def set_output_type(self, output_id: str):
        """Set the output type by ID."""
        for display_name, id_ in OUTPUT_TYPE_DISPLAY_TO_ID.items():
            if id_ == output_id:
                self.output_type_combo.set(display_name)
                self._on_selection_change()
                break


# =============================================================================
# INNOVATION FORM COMPONENT (WITH EXPORT/IMPORT & OUTPUT-SPECIFIC FIELDS)
# =============================================================================

class InnovationFormComponent(tk.Frame):
    """
    Enhanced form for entering innovation details.
    Includes Excel import/export functionality, expanded fields, vertical scrolling,
    and "Other" option support for all dropdowns.
    Supports output-specific field export/import.
    """
    
    def __init__(self, parent, on_generate: Callable, on_clear: Callable, 
                 on_import: Optional[Callable] = None):
        super().__init__(parent, bg=ADI_COLORS["white"])
        self.on_generate = on_generate
        self.on_clear = on_clear
        self.on_import = on_import
        
        # Track "Other" text fields for each combobox
        self.other_entries: Dict[str, tk.Entry] = {}
        self.other_frames: Dict[str, tk.Frame] = {}
        
        # Reference to output-specific inputs component (set externally)
        self.output_specific_component = None
        self.current_output_type = "excel_value_model"
        
        self._create_widgets()
    
    def set_output_specific_component(self, component):
        """Set reference to the output-specific inputs component for export/import."""
        self.output_specific_component = component
    
    def _create_widgets(self):
        # ===== TOP BAR (Fixed - doesn't scroll) =====
        top_bar = tk.Frame(self, bg=ADI_COLORS["white"])
        top_bar.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # Title row with import/export buttons
        title_row = tk.Frame(top_bar, bg=ADI_COLORS["white"])
        title_row.pack(fill=tk.X)
        
        tk.Label(
            title_row,
            text="Innovation Details",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["dark_gray"]
        ).pack(side=tk.LEFT)
        
        # Import from Excel button
        import_btn = tk.Button(
            title_row,
            text="📥 Import from Excel",
            command=self._handle_import,
            bg=ADI_COLORS["primary_blue"],
            fg="black",
            font=(UI_CONFIG["font_family"], 10, "bold"),
            padx=15,
            pady=3,
            cursor="hand2"
        )
        import_btn.pack(side=tk.RIGHT)
        
        # Export to Excel button
        export_btn = tk.Button(
            title_row,
            text="📤 Export to Excel",
            command=self._export_to_excel,
            bg=ADI_COLORS["success_green"],
            fg="black",
            font=(UI_CONFIG["font_family"], 10, "bold"),
            padx=15,
            pady=3,
            cursor="hand2"
        )
        export_btn.pack(side=tk.RIGHT, padx=10)
        
        # Download template button
        template_btn = tk.Button(
            title_row,
            text="📄 Download Template",
            command=self._download_template,
            bg=ADI_COLORS["medium_gray"],
            fg="black",
            font=(UI_CONFIG["font_family"], 9),
            padx=10,
            pady=3,
            cursor="hand2"
        )
        template_btn.pack(side=tk.RIGHT, padx=10)
        
        tk.Label(
            top_bar,
            text="Enter details manually or import from Excel file. Scroll down for more fields.",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["text_gray"]
        ).pack(anchor="w", pady=(5, 0))
        
        # ===== SCROLLABLE FORM AREA =====
        self.scrollable = ScrollableFrame(
            self, 
            bg_color=ADI_COLORS["white"],
            always_show_vertical=True
        )
        self.scrollable.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Get the inner frame for adding content
        form_container = self.scrollable.get_frame()
        
        # Create all form fields inside the scrollable area
        self._create_form_fields(form_container)
        
        # ===== BOTTOM BAR (Fixed - doesn't scroll) =====
        bottom_bar = tk.Frame(self, bg=ADI_COLORS["white"])
        bottom_bar.pack(fill=tk.X, padx=10, pady=10)
        
        # Separator
        ttk.Separator(bottom_bar, orient="horizontal").pack(fill=tk.X, pady=(0, 10))
        
        # Buttons
        btn_frame = tk.Frame(bottom_bar, bg=ADI_COLORS["white"])
        btn_frame.pack(fill=tk.X)
        
        tk.Button(
            btn_frame,
            text="Clear Form",
            command=self._clear_and_scroll_top,
            bg=ADI_COLORS["medium_gray"],
            fg="black",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            padx=15,
            pady=5,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=15)
        
        # Scroll buttons
        scroll_btn_frame = tk.Frame(btn_frame, bg=ADI_COLORS["white"])
        scroll_btn_frame.pack(side=tk.RIGHT)
        
        tk.Button(
            scroll_btn_frame,
            text="⬆ Top",
            command=self._scroll_to_top,
            bg=ADI_COLORS["light_gray"],
            fg=ADI_COLORS["dark_gray"],
            font=(UI_CONFIG["font_family"], 9),
            padx=8,
            pady=3,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=2)
        
        tk.Button(
            scroll_btn_frame,
            text="⬇ Bottom",
            command=self._scroll_to_bottom,
            bg=ADI_COLORS["light_gray"],
            fg=ADI_COLORS["dark_gray"],
            font=(UI_CONFIG["font_family"], 9),
            padx=8,
            pady=3,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=2)
    
    # =========================================================================
    # EXPORT TO EXCEL - WITH OUTPUT-SPECIFIC FIELDS
    # =========================================================================
    
    def _export_to_excel(self):
        """Export current form values including output-specific fields to an Excel file."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        
        # Get all form values
        values = self.get_values()
        
        # Get output-specific values if component is available
        specific_values = {}
        if self.output_specific_component:
            specific_values = self.output_specific_component.get_values()
        
        # Check if there's any data to export
        if not any(values.values()) and not any(specific_values.values()):
            messagebox.showwarning(
                "No Data",
                "Please fill in at least one field before exporting."
            )
            return
        
        # Generate default filename
        innovation_name = values.get("innovation_name", "Innovation")
        safe_name = "".join(c for c in innovation_name if c.isalnum() or c in (' ', '-', '_')).rstrip()[:30]
        safe_name = safe_name.replace(' ', '_') if safe_name else "Innovation"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{safe_name}_FormData_{timestamp}.xlsx"
        
        # Prompt for save location
        filepath = filedialog.asksaveasfilename(
            title="Export Form Data to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not filepath:
            return  # User cancelled
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Innovation Form Data"
            
            # Define styles
            header_fill = PatternFill(start_color="0067B9", end_color="0067B9", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="4A9BD9", end_color="4A9BD9", fill_type="solid")
            section_font = Font(bold=True, color="FFFFFF", size=11)
            field_font = Font(bold=True, size=11)
            value_font = Font(size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            # Title row
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = f"Innovation Form Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            title_cell.font = Font(bold=True, size=14, color="0067B9")
            title_cell.alignment = Alignment(horizontal='center')
            
            # Header row
            ws['A3'] = "Field"
            ws['B3'] = "Value"
            ws['C3'] = "Notes"
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}3']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # ===== BASE FORM FIELDS =====
            row = 4
            
            # Section header: Required Fields
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value="--- REQUIRED FIELDS ---")
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=1).font = section_font
            row += 1
            
            base_fields = [
                ("Innovation Name", "innovation_name", "Required field"),
                ("Innovation Description", "innovation_description", "Required field - 2-4 sentences"),
                ("Target Industry", "industry", "Required field"),
            ]
            
            for label, key, note in base_fields:
                ws.cell(row=row, column=1, value=label).font = field_font
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2, value=values.get(key, "")).font = value_font
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=2).alignment = wrap_alignment
                ws.cell(row=row, column=3, value=note).font = Font(size=10, italic=True, color="666666")
                ws.cell(row=row, column=3).border = thin_border
                row += 1
            
            # Section header: Analysis Parameters
            row += 1
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value="--- ANALYSIS PARAMETERS ---")
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=1).font = section_font
            row += 1
            
            param_fields = [
                ("Geographic Scope", "geographic_scope", "Options: United States, North America, Europe, Asia-Pacific, Global, or custom"),
                ("Analysis Timeframe", "analysis_timeframe", "Options: Year 1 at Scale, 3-Year, 5-Year, 10-Year Projection, or custom"),
                ("Innovation Stage", "innovation_stage", "Options: Concept, Prototype, Pilot, Commercial, Growth, or custom"),
                ("Currency", "currency", "Options: USD, EUR, GBP, JPY, or custom"),
            ]
            
            for label, key, note in param_fields:
                ws.cell(row=row, column=1, value=label).font = field_font
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2, value=values.get(key, "")).font = value_font
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=3, value=note).font = Font(size=10, italic=True, color="666666")
                ws.cell(row=row, column=3).border = thin_border
                row += 1
            
            # Add Output Type
            ws.cell(row=row, column=1, value="Output Type").font = field_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=self.current_output_type).font = value_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value="excel_value_model, powerpoint_pitch, or word_gonogo")
            ws.cell(row=row, column=3).font = Font(size=10, italic=True, color="666666")
            ws.cell(row=row, column=3).border = thin_border
            row += 1
            
            # ===== OUTPUT-SPECIFIC FIELDS =====
            if specific_values and any(specific_values.values()):
                row += 1
                
                # Determine output type name
                output_type_name = self.current_output_type.replace("_", " ").title()
                
                ws.merge_cells(f'A{row}:C{row}')
                ws.cell(row=row, column=1, value=f"--- OUTPUT-SPECIFIC FIELDS ({output_type_name}) ---")
                ws.cell(row=row, column=1).fill = section_fill
                ws.cell(row=row, column=1).font = section_font
                row += 1
                
                # Get field labels based on output type
                specific_field_labels = self._get_specific_field_labels()
                
                for key, value in specific_values.items():
                    if value:  # Only export non-empty fields
                        label = specific_field_labels.get(key, key.replace("_", " ").title())
                        
                        ws.cell(row=row, column=1, value=label).font = field_font
                        ws.cell(row=row, column=1).border = thin_border
                        ws.cell(row=row, column=2, value=value).font = value_font
                        ws.cell(row=row, column=2).border = thin_border
                        ws.cell(row=row, column=2).alignment = wrap_alignment
                        ws.cell(row=row, column=3, value=f"Output-specific: {self.current_output_type}")
                        ws.cell(row=row, column=3).font = Font(size=10, italic=True, color="666666")
                        ws.cell(row=row, column=3).border = thin_border
                        row += 1
            
            # ===== OTHER FIELD DETAILS =====
            # Get "Other" field values
            other_has_values = False
            for field_name, entry in self.other_entries.items():
                if entry.get().strip():
                    other_has_values = True
                    break
            
            if other_has_values:
                row += 1
                ws.merge_cells(f'A{row}:C{row}')
                ws.cell(row=row, column=1, value="--- OTHER FIELD DETAILS ---")
                ws.cell(row=row, column=1).fill = section_fill
                ws.cell(row=row, column=1).font = section_font
                row += 1
                
                other_field_labels = {
                    "geographic_scope": "Geographic Scope (Other Detail)",
                    "analysis_timeframe": "Analysis Timeframe (Other Detail)",
                    "innovation_stage": "Innovation Stage (Other Detail)",
                    "currency": "Currency (Other Detail)",
                }
                
                for field_name, entry in self.other_entries.items():
                    other_value = entry.get().strip()
                    if other_value:
                        label = other_field_labels.get(field_name, f"{field_name} (Other)")
                        ws.cell(row=row, column=1, value=label).font = field_font
                        ws.cell(row=row, column=1).border = thin_border
                        ws.cell(row=row, column=2, value=other_value).font = value_font
                        ws.cell(row=row, column=2).border = thin_border
                        ws.cell(row=row, column=3, value="Custom value for 'Other' selection")
                        ws.cell(row=row, column=3).font = Font(size=10, italic=True, color="666666")
                        ws.cell(row=row, column=3).border = thin_border
                        row += 1
            
            # ===== METADATA SECTION =====
            row += 2
            ws.cell(row=row, column=1, value="Export Metadata")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="0067B9")
            row += 1
            
            ws.cell(row=row, column=1, value="Export Date/Time")
            ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            row += 1
            
            ws.cell(row=row, column=1, value="Output Type")
            ws.cell(row=row, column=2, value=self.current_output_type)
            row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 55
            
            # Save the workbook
            wb.save(filepath)
            
            # Count exported fields
            base_count = len([v for v in values.values() if v])
            specific_count = len([v for v in specific_values.values() if v]) if specific_values else 0
            
            messagebox.showinfo(
                "Export Successful",
                f"Form data exported successfully!\n\n"
                f"• {base_count} base fields\n"
                f"• {specific_count} output-specific fields\n\n"
                f"File: {filepath}"
            )
            
        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Failed to export form data:\n{str(e)}"
            )
    
    def _get_specific_field_labels(self) -> Dict[str, str]:
        """Get human-readable labels for output-specific fields."""
        labels = {
            # Excel Value Model fields
            "value_drivers": "Primary Value Drivers",
            "value_factors": "Value Calculation Factors",
            "stakeholders": "Stakeholder Identification",
            "value_allocation": "Value Allocation Guidance",
            "segments_include": "Segments to Include",
            "segments_exclude": "Segments to Exclude",
            "growth_assumptions": "Growth Rate Guidance",
            
            # PowerPoint Pitch fields
            "problem_primary": "Primary Problem Statement",
            "problem_secondary": "Supporting Problems / Pain Points",
            "problem_stats": "Key Statistics & Evidence",
            "solution_how": "How It Works",
            "solution_features": "Key Features & Capabilities",
            "solution_benefits": "Quantified Benefits",
            "value_prop": "Core Value Proposition",
            "differentiators": "Key Differentiators",
            "the_ask": "Investment/Resource Request",
            "use_of_funds": "Use of Funds / Resources",
            
            # Word Deep Dive fields
            "exec_summary": "Executive Summary Guidance",
            "purpose_scope": "Purpose & Scope",
            "tech_market": "Technology & Market Overview",
            "competitive": "Competitive Landscape",
            "rationale": "Rationale for Go Decision",
            "risks": "Risks & Unknowns",
            "triggers": "Technology Triggers",
            "collaboration": "Collaboration Opportunities",
            "next_steps": "Next Steps",
        }
        return labels
    
    # =========================================================================
    # FORM FIELD CREATION
    # =========================================================================
    
    def _create_form_fields(self, parent):
        """Create all form fields organized in sections."""
        
        # ===== REQUIRED FIELDS SECTION =====
        self._create_section_header(parent, "Required Fields", is_first=True)
        
        form_frame = tk.Frame(parent, bg=ADI_COLORS["white"])
        form_frame.pack(fill=tk.X, padx=10)
        
        # Row 0: Innovation Name
        self._create_label(form_frame, "Innovation Name *", 0, 0)
        self.name_entry = self._create_entry(form_frame, 0, 1, width=55)
        
        # Row 1: Target Industry
        self._create_label(form_frame, "Target Industry *", 1, 0)
        self.industry_entry = self._create_entry(form_frame, 1, 1, width=45)
        
        # Description (multi-line)
        desc_frame = tk.Frame(parent, bg=ADI_COLORS["white"])
        desc_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
        
        tk.Label(
            desc_frame,
            text="Innovation Description *",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"], "bold"),
            bg=ADI_COLORS["white"]
        ).pack(anchor="w")
        
        tk.Label(
            desc_frame,
            text="Describe what it does, how it works, and the problem it solves (2-4 sentences)",
            font=(UI_CONFIG["font_family"], 9),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["text_gray"]
        ).pack(anchor="w")
        
        self.desc_text = tk.Text(
            desc_frame,
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            height=4,
            wrap=tk.WORD,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            padx=5,
            pady=5
        )
        self.desc_text.pack(fill=tk.X, pady=5)
        
        # ===== ANALYSIS PARAMETERS SECTION =====
        self._create_section_header(parent, "Analysis Parameters")
        
        params_frame = tk.Frame(parent, bg=ADI_COLORS["white"])
        params_frame.pack(fill=tk.X, padx=10)
        
        # Row 0: Geographic Scope with inline "Other"
        self._create_label(params_frame, "Geographic Scope", 0, 0)
        geo_container = tk.Frame(params_frame, bg=ADI_COLORS["white"])
        geo_container.grid(row=0, column=1, sticky="w", padx=(0, 15), pady=5, columnspan=3)
        
        self.geo_combo = ttk.Combobox(
            geo_container, 
            width=18, 
            state="readonly", 
            values=FORM_OPTIONS["geographic_scope"]
        )
        self.geo_combo.set(FORM_DEFAULTS["geographic_scope"])
        self.geo_combo.pack(side=tk.LEFT)
        self.geo_combo.bind("<<ComboboxSelected>>", 
                           lambda e: self._on_combo_change("geographic_scope"))
        
        self.geo_other_entry = tk.Entry(
            geo_container,
            font=(UI_CONFIG["font_family"], 10),
            width=25,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            bd=3
        )
        self.other_entries["geographic_scope"] = self.geo_other_entry
        
        # Row 1: Analysis Timeframe with inline "Other"
        self._create_label(params_frame, "Analysis Timeframe", 1, 0)
        time_container = tk.Frame(params_frame, bg=ADI_COLORS["white"])
        time_container.grid(row=1, column=1, sticky="w", padx=(0, 15), pady=5, columnspan=3)
        
        self.time_combo = ttk.Combobox(
            time_container, 
            width=18, 
            state="readonly", 
            values=FORM_OPTIONS["analysis_timeframe"]
        )
        self.time_combo.set(FORM_DEFAULTS["analysis_timeframe"])
        self.time_combo.pack(side=tk.LEFT)
        self.time_combo.bind("<<ComboboxSelected>>", 
                            lambda e: self._on_combo_change("analysis_timeframe"))
        
        self.time_other_entry = tk.Entry(
            time_container,
            font=(UI_CONFIG["font_family"], 10),
            width=25,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            bd=3
        )
        self.other_entries["analysis_timeframe"] = self.time_other_entry
        
        # Row 2: Innovation Stage with inline "Other"
        self._create_label(params_frame, "Innovation Stage", 2, 0)
        stage_container = tk.Frame(params_frame, bg=ADI_COLORS["white"])
        stage_container.grid(row=2, column=1, sticky="w", padx=(0, 15), pady=5, columnspan=3)
        
        self.stage_combo = ttk.Combobox(
            stage_container, 
            width=14, 
            state="readonly", 
            values=FORM_OPTIONS["innovation_stage"]
        )
        self.stage_combo.set(FORM_DEFAULTS["innovation_stage"])
        self.stage_combo.pack(side=tk.LEFT)
        self.stage_combo.bind("<<ComboboxSelected>>", 
                             lambda e: self._on_combo_change("innovation_stage"))
        
        self.stage_other_entry = tk.Entry(
            stage_container,
            font=(UI_CONFIG["font_family"], 10),
            width=20,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            bd=3
        )
        self.other_entries["innovation_stage"] = self.stage_other_entry
        
        # Row 3: Currency with inline "Other"
        self._create_label(params_frame, "Currency", 3, 0)
        currency_container = tk.Frame(params_frame, bg=ADI_COLORS["white"])
        currency_container.grid(row=3, column=1, sticky="w", padx=(0, 15), pady=5, columnspan=3)
        
        self.currency_combo = ttk.Combobox(
            currency_container, 
            width=8, 
            state="readonly", 
            values=FORM_OPTIONS["currency"]
        )
        self.currency_combo.set(FORM_DEFAULTS["currency"])
        self.currency_combo.pack(side=tk.LEFT)
        self.currency_combo.bind("<<ComboboxSelected>>", 
                                lambda e: self._on_combo_change("currency"))
        
        self.currency_other_entry = tk.Entry(
            currency_container,
            font=(UI_CONFIG["font_family"], 10),
            width=15,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            bd=3
        )
        self.other_entries["currency"] = self.currency_other_entry
        
        # Add padding at the bottom
        spacer = tk.Frame(parent, bg=ADI_COLORS["white"], height=20)
        spacer.pack(fill=tk.X)
    
    def _on_combo_change(self, field_name: str):
        """Handle combobox selection change - show/hide inline 'Other' field."""
        combo_map = {
            "geographic_scope": (self.geo_combo, self.geo_other_entry),
            "analysis_timeframe": (self.time_combo, self.time_other_entry),
            "innovation_stage": (self.stage_combo, self.stage_other_entry),
            "currency": (self.currency_combo, self.currency_other_entry),
        }
        
        combo, other_entry = combo_map.get(field_name, (None, None))
        if not combo or not other_entry:
            return
        
        selected_value = combo.get()
        
        if selected_value == "Other":
            if not other_entry.winfo_ismapped():
                other_entry.pack(side=tk.LEFT, padx=(10, 0))
                other_entry.focus_set()
        else:
            if other_entry.winfo_ismapped():
                other_entry.pack_forget()
            other_entry.delete(0, tk.END)
    
    def _create_section_header(self, parent, text: str, is_first: bool = False):
        """Create a section header with separator line."""
        frame = tk.Frame(parent, bg=ADI_COLORS["white"])
        frame.pack(fill=tk.X, padx=10, pady=(5 if is_first else 15, 5))
        
        tk.Label(
            frame,
            text=text,
            font=(UI_CONFIG["font_family"], 11, "bold"),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["primary_blue"]
        ).pack(side=tk.LEFT)
        
        sep_container = tk.Frame(frame, bg=ADI_COLORS["white"])
        sep_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        ttk.Separator(sep_container, orient="horizontal").pack(fill=tk.X, pady=8)
    
    def _create_label(self, parent, text: str, row: int, col: int, bg=None):
        if bg is None:
            bg = ADI_COLORS["white"]
        label = tk.Label(
            parent,
            text=text,
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"], "bold"),
            bg=bg
        )
        label.grid(row=row, column=col, sticky="w", pady=5, padx=(0, 5))
        return label
    
    def _create_entry(self, parent, row: int, col: int, width: int = 40, bg=None):
        if bg is None:
            bg = ADI_COLORS["light_gray"]
        entry = tk.Entry(
            parent,
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            width=width,
            bg=bg,
            relief=tk.FLAT,
            bd=5
        )
        entry.grid(row=row, column=col, sticky="w", padx=(0, 15), pady=5)
        return entry
    
    # =========================================================================
    # IMPORT/EXPORT HANDLERS
    # =========================================================================
    
    def _handle_import(self):
        """Handle import from Excel button click."""
        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if filepath and self.on_import:
            self.on_import(filepath)
    
    def _download_template(self):
        """Download/create Excel template for current output type."""
        filepath = filedialog.asksaveasfilename(
            title="Save Template As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Innovation_Input_Template_{self.current_output_type}.xlsx"
        )
        
        if filepath:
            try:
                from excel_importer import excel_importer
                if excel_importer.create_template(filepath, self.current_output_type):
                    messagebox.showinfo(
                        "Template Created",
                        f"Template saved to:\n{filepath}\n\n"
                        f"Output type: {self.current_output_type}\n\n"
                        "Fill in the values and use 'Import from Excel' to load."
                    )
                else:
                    messagebox.showerror("Error", "Failed to create template.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create template:\n{str(e)}")
    
    def _scroll_to_top(self):
        """Scroll the form to the top."""
        self.scrollable.scroll_to_top()
    
    def _scroll_to_bottom(self):
        """Scroll the form to the bottom."""
        self.scrollable.scroll_to_bottom()
    
    def _clear_and_scroll_top(self):
        """Clear the form and scroll to top."""
        self.clear()
        self.scrollable.scroll_to_top()
        if self.on_clear:
            self.on_clear()
    
    # =========================================================================
    # POPULATE FROM IMPORTED DATA
    # =========================================================================
    
    def populate_from_data(self, data: Dict[str, str]):
        """
        Populate form fields from imported data.
        Handles values that don't match dropdown options by selecting 'Other'.
        """
        self.clear()
        self.scrollable.scroll_to_top()
        
        # Required fields
        if "innovation_name" in data:
            self.name_entry.insert(0, data["innovation_name"])
        
        if "industry" in data:
            self.industry_entry.insert(0, data["industry"])
        
        if "innovation_description" in data:
            self.desc_text.insert("1.0", data["innovation_description"])
        
        # Analysis parameters with "Other" handling
        if "geographic_scope" in data:
            self._set_combo_value_with_other(
                self.geo_combo, 
                self.geo_other_entry,
                data["geographic_scope"]
            )
        
        if "analysis_timeframe" in data:
            self._set_combo_value_with_other(
                self.time_combo, 
                self.time_other_entry,
                data["analysis_timeframe"]
            )
        
        if "innovation_stage" in data:
            self._set_combo_value_with_other(
                self.stage_combo, 
                self.stage_other_entry,
                data["innovation_stage"]
            )
        
        if "currency" in data:
            self._set_combo_value_with_other(
                self.currency_combo, 
                self.currency_other_entry,
                data["currency"]
            )
        
        # Set output type if provided
        if "output_type" in data:
            self.current_output_type = data["output_type"]
    
    def populate_specific_fields(self, data: Dict[str, str]):
        """Populate output-specific fields from imported data."""
        if self.output_specific_component and data:
            current_type = self.current_output_type
            if current_type in self.output_specific_component.form_widgets:
                widgets = self.output_specific_component.form_widgets[current_type]
                for key, widget in widgets.items():
                    if key in data and data[key]:
                        if isinstance(widget, tk.Text):
                            widget.delete("1.0", tk.END)
                            widget.insert("1.0", data[key])
                        elif isinstance(widget, tk.Entry):
                            widget.delete(0, tk.END)
                            widget.insert(0, data[key])
    
    def _set_combo_value_with_other(self, combo: ttk.Combobox, other_entry: tk.Entry, value: str):
        """
        Set combobox value, selecting 'Other' and populating the inline text field
        if the value doesn't match any dropdown option.
        """
        values = list(combo["values"])
        value_stripped = value.strip()
        value_lower = value_stripped.lower()
        
        # First, try exact match (case-insensitive)
        for v in values:
            if v.lower() == value_lower and v != "Other":
                combo.set(v)
                if other_entry.winfo_ismapped():
                    other_entry.pack_forget()
                other_entry.delete(0, tk.END)
                return
        
        # Second, try partial match
        for v in values:
            if v != "Other" and (value_lower in v.lower() or v.lower() in value_lower):
                combo.set(v)
                if other_entry.winfo_ismapped():
                    other_entry.pack_forget()
                other_entry.delete(0, tk.END)
                return
        
        # No match found - select "Other" and populate the inline text field
        if "Other" in values:
            combo.set("Other")
            other_entry.delete(0, tk.END)
            other_entry.insert(0, value_stripped)
            if not other_entry.winfo_ismapped():
                other_entry.pack(side=tk.LEFT, padx=(10, 0))
        else:
            if values:
                combo.set(values[0])
    
    # =========================================================================
    # GET/CLEAR VALUES
    # =========================================================================
    
    def get_values(self) -> Dict[str, str]:
        """
        Get all form values including extended fields.
        Returns the 'Other' text field value when 'Other' is selected.
        """
        # Get base dropdown values, replacing "Other" with actual text
        geo_value = self.geo_combo.get()
        if geo_value == "Other":
            other_text = self.geo_other_entry.get().strip()
            geo_value = other_text if other_text else "Other (unspecified)"
        
        time_value = self.time_combo.get()
        if time_value == "Other":
            other_text = self.time_other_entry.get().strip()
            time_value = other_text if other_text else "Other (unspecified)"
        
        stage_value = self.stage_combo.get()
        if stage_value == "Other":
            other_text = self.stage_other_entry.get().strip()
            stage_value = other_text if other_text else "Other (unspecified)"
        
        currency_value = self.currency_combo.get()
        if currency_value == "Other":
            other_text = self.currency_other_entry.get().strip()
            currency_value = other_text if other_text else "Other (unspecified)"
        
        values = {
            "innovation_name": self.name_entry.get().strip(),
            "innovation_description": self.desc_text.get("1.0", tk.END).strip(),
            "industry": self.industry_entry.get().strip(),
            "geographic_scope": geo_value,
            "analysis_timeframe": time_value,
            "innovation_stage": stage_value,
            "currency": currency_value,
        }
        
        return values
    
    def clear(self):
        """Clear all form fields."""
        self.name_entry.delete(0, tk.END)
        self.industry_entry.delete(0, tk.END)
        self.desc_text.delete("1.0", tk.END)
        
        # Reset comboboxes to defaults
        self.geo_combo.set(FORM_DEFAULTS["geographic_scope"])
        self.time_combo.set(FORM_DEFAULTS["analysis_timeframe"])
        self.stage_combo.set(FORM_DEFAULTS["innovation_stage"])
        self.currency_combo.set(FORM_DEFAULTS["currency"])
        
        # Clear and hide all inline "Other" entry fields
        for field_name, entry in self.other_entries.items():
            entry.delete(0, tk.END)
            if entry.winfo_ismapped():
                entry.pack_forget()




# =============================================================================
# OUTPUT-TYPE-SPECIFIC FORM COMPONENT (WITH SCROLLING)
# =============================================================================

class OutputSpecificInputsComponent(tk.Frame):
    """
    Dynamic form component that shows different inputs based on selected output type.
    Each form type has its own scrollable section.
    """
    
    def __init__(self, parent, bg_color: str = None):
        super().__init__(parent, bg=bg_color or ADI_COLORS["white"])
        self.bg_color = bg_color or ADI_COLORS["white"]
        self.current_output_type = "excel_value_model"
        
        self.form_frames: Dict[str, tk.Frame] = {}
        self.form_widgets: Dict[str, Dict[str, Any]] = {}
        
        self._create_all_forms()
        self._show_form("excel_value_model")
    
    def _create_all_forms(self):
        """Create all output-type-specific forms (initially hidden)."""
        self._create_excel_form()
        self._create_deepdive_form()
        self._create_powerpoint_form()
    
    def _create_section_header(self, parent, title: str, subtitle: str = "", bg: str = None):
        """Create a styled section header."""
        bg = bg or self.bg_color
        header_frame = tk.Frame(parent, bg=bg)
        header_frame.pack(fill=tk.X, pady=(10, 5))
        
        tk.Label(
            header_frame,
            text=title,
            font=(UI_CONFIG["font_family"], 11, "bold"),
            bg=bg,
            fg=ADI_COLORS["primary_blue"]
        ).pack(anchor="w")
        
        if subtitle:
            tk.Label(
                header_frame,
                text=subtitle,
                font=(UI_CONFIG["font_family"], 9),
                bg=bg,
                fg=ADI_COLORS["text_gray"],
                wraplength=700
            ).pack(anchor="w")
    
    def _create_text_input(self, parent, label: str, hint: str, height: int = 3, bg: str = None) -> tk.Text:
        """Create a labeled text input with hint."""
        bg = bg or self.bg_color
        frame = tk.Frame(parent, bg=bg)
        frame.pack(fill=tk.X, pady=5, padx=5)
        
        tk.Label(
            frame,
            text=label,
            font=(UI_CONFIG["font_family"], 10, "bold"),
            bg=bg,
            fg=ADI_COLORS["dark_gray"]
        ).pack(anchor="w")
        
        tk.Label(
            frame,
            text=hint,
            font=(UI_CONFIG["font_family"], 9),
            bg=bg,
            fg=ADI_COLORS["text_gray"],
            wraplength=700
        ).pack(anchor="w")
        
        text_widget = tk.Text(
            frame,
            font=(UI_CONFIG["font_family"], 10),
            height=height,
            wrap=tk.WORD,
            bg=ADI_COLORS["light_gray"],
            relief=tk.FLAT,
            padx=8,
            pady=6
        )
        text_widget.pack(fill=tk.X, pady=(3, 0))
        
        return text_widget
    
    # =========================================================================
    # EXCEL VALUE MODEL FORM (WITH SCROLLING)
    # =========================================================================
    
    def _create_excel_form(self):
        """Create form for Excel Value Creation Model inputs."""
        frame = tk.Frame(self, bg=self.bg_color)
        self.form_frames["excel_value_model"] = frame
        self.form_widgets["excel_value_model"] = {}
        
        # Header
        header = tk.Frame(frame, bg=ADI_COLORS["success_green"], pady=8)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="📊 Value Creation Model - Specific Inputs",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["success_green"],
            fg="white"
        ).pack()
        
        # Scrollable content
        scroll_container = ScrollableFrame(
            frame, 
            bg_color=self.bg_color,
            always_show_vertical=True
        )
        scroll_container.pack(fill=tk.BOTH, expand=True)
        content = scroll_container.get_frame()
        
        # Section 1: Value Drivers
        self._create_section_header(
            content,
            "💰 Value Drivers",
            "What are the main drivers of value creation? What factors should be considered?"
        )
        
        self.form_widgets["excel_value_model"]["value_drivers"] = self._create_text_input(
            content,
            "Primary Value Drivers",
            "List 3-5 main ways this innovation creates value (e.g., cost reduction, revenue increase, risk mitigation)",
            height=4
        )
        
        self.form_widgets["excel_value_model"]["value_factors"] = self._create_text_input(
            content,
            "Value Calculation Factors",
            "What factors/metrics should be used to quantify each driver?",
            height=3
        )
        
        # Section 2: Stakeholders
        self._create_section_header(
            content,
            "👥 Key Stakeholders",
            "Who are the stakeholders that will derive value from this innovation?"
        )
        
        self.form_widgets["excel_value_model"]["stakeholders"] = self._create_text_input(
            content,
            "Stakeholder Identification",
            "List key stakeholders and briefly describe how each captures value",
            height=4
        )
        
        self.form_widgets["excel_value_model"]["value_allocation"] = self._create_text_input(
            content,
            "Value Allocation Guidance",
            "How should value be allocated among stakeholders?",
            height=2
        )
        
        # Section 3: Market Segments
        self._create_section_header(
            content,
            "🎯 Market Segmentation",
            "Which market segments should be included or excluded?"
        )
        
        self.form_widgets["excel_value_model"]["segments_include"] = self._create_text_input(
            content,
            "Segments to Include",
            "List specific market segments to analyze",
            height=3
        )
        
        self.form_widgets["excel_value_model"]["segments_exclude"] = self._create_text_input(
            content,
            "Segments to Exclude (and why)",
            "List any segments to explicitly exclude and rationale",
            height=2
        )
        
        # Section 4: Growth Assumptions
        self._create_section_header(
            content,
            "📈 Growth & Projection Guidance",
            "Any specific assumptions for growth rates and projections?"
        )
        
        self.form_widgets["excel_value_model"]["growth_assumptions"] = self._create_text_input(
            content,
            "Growth Rate Guidance",
            "Expected growth drivers, market expansion factors, or constraints",
            height=3
        )
        
        # Spacer
        tk.Frame(content, bg=self.bg_color, height=30).pack(fill=tk.X)
    
    # =========================================================================
    # DEEP DIVE SUMMARY FORM (WITH SCROLLING)
    # =========================================================================
    
    def _create_deepdive_form(self):
        """Create form for Deep Dive Summary inputs - section by section."""
        frame = tk.Frame(self, bg=self.bg_color)
        self.form_frames["word_gonogo"] = frame
        self.form_widgets["word_gonogo"] = {}
        
        # Header
        header = tk.Frame(frame, bg=ADI_COLORS["primary_blue"], pady=8)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="📋 Deep Dive Summary - Section Content Guidance",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["primary_blue"],
            fg="white"
        ).pack()
        
        tk.Label(
            header,
            text="Provide 2-3 sentences describing the proposed content for each section",
            font=(UI_CONFIG["font_family"], 10),
            bg=ADI_COLORS["primary_blue"],
            fg=ADI_COLORS["light_blue"]
        ).pack()
        
        # Scrollable content
        scroll_container = ScrollableFrame(
            frame, 
            bg_color=self.bg_color,
            always_show_vertical=True
        )
        scroll_container.pack(fill=tk.BOTH, expand=True)
        content = scroll_container.get_frame()
        
        # Executive Summary
        self._create_section_header(content, "1️⃣ Executive Summary")
        self.form_widgets["word_gonogo"]["exec_summary"] = self._create_text_input(
            content,
            "Brief Overview & Decision Statement",
            "What is the key finding and recommendation?",
            height=3
        )
        
        # Purpose & Scope
        self._create_section_header(content, "2️⃣ Purpose & Scope")
        self.form_widgets["word_gonogo"]["purpose_scope"] = self._create_text_input(
            content,
            "Objective & Exploration Focus",
            "What was assessed? What questions will Exploration address?",
            height=3
        )
        
        # Technology & Market
        self._create_section_header(content, "3️⃣ Current State of Technology & Market")
        self.form_widgets["word_gonogo"]["tech_market"] = self._create_text_input(
            content,
            "Technology & Market Overview",
            "Key technologies, maturity levels, market size, customer segments",
            height=3
        )
        
        # Competitive Landscape
        self._create_section_header(content, "4️⃣ Competitive Landscape")
        self.form_widgets["word_gonogo"]["competitive"] = self._create_text_input(
            content,
            "Competitors & Positioning",
            "Key competitors, their offerings, ADI's competitive position",
            height=3
        )
        
        # Rationale for Go
        self._create_section_header(content, "5️⃣ Rationale for Go Decision")
        self.form_widgets["word_gonogo"]["rationale"] = self._create_text_input(
            content,
            "Key Findings & Reasons to Proceed",
            "Most important findings supporting the recommendation",
            height=3
        )
        
        # Risks & Unknowns
        self._create_section_header(content, "6️⃣ Risks & Unknowns")
        self.form_widgets["word_gonogo"]["risks"] = self._create_text_input(
            content,
            "Key Risks & Mitigation",
            "Major risks, IP strategy considerations",
            height=3
        )
        
        # Technology Triggers
        self._create_section_header(content, "7️⃣ Technology Triggers & Market Dynamics")
        self.form_widgets["word_gonogo"]["triggers"] = self._create_text_input(
            content,
            "External Factors",
            "Breakthroughs, standards, regulatory shifts",
            height=2
        )
        
        # Collaboration
        self._create_section_header(content, "8️⃣ Collaboration & Companies to Watch")
        self.form_widgets["word_gonogo"]["collaboration"] = self._create_text_input(
            content,
            "Partners & Monitoring",
            "Potential partners, companies and trends to monitor",
            height=3
        )
        
        # Next Steps
        self._create_section_header(content, "9️⃣ Next Steps")
        self.form_widgets["word_gonogo"]["next_steps"] = self._create_text_input(
            content,
            "Immediate Actions",
            "What needs to happen to launch Exploration?",
            height=2
        )
        
        # Spacer
        tk.Frame(content, bg=self.bg_color, height=30).pack(fill=tk.X)
    
    # =========================================================================
    # POWERPOINT PITCH FORM (WITH SCROLLING)
    # =========================================================================
    
    def _create_powerpoint_form(self):
        """Create form for PowerPoint Pitch Deck inputs."""
        frame = tk.Frame(self, bg=self.bg_color)
        self.form_frames["powerpoint_pitch"] = frame
        self.form_widgets["powerpoint_pitch"] = {}
        
        # Header
        header = tk.Frame(frame, bg=ADI_COLORS["accent_orange"], pady=8)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="📽️ Executive Pitch Deck - Key Messaging Inputs",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["accent_orange"],
            fg="white"
        ).pack()
        
        # Scrollable content
        scroll_container = ScrollableFrame(
            frame, 
            bg_color=self.bg_color,
            always_show_vertical=True
        )
        scroll_container.pack(fill=tk.BOTH, expand=True)
        content = scroll_container.get_frame()
        
        # Section 1: Problem Statements
        self._create_section_header(
            content,
            "🔴 Problem Statements",
            "What are the key problems this innovation addresses?"
        )
        
        self.form_widgets["powerpoint_pitch"]["problem_primary"] = self._create_text_input(
            content,
            "Primary Problem Statement",
            "The #1 problem this solves - make it compelling and specific",
            height=3
        )
        
        self.form_widgets["powerpoint_pitch"]["problem_secondary"] = self._create_text_input(
            content,
            "Supporting Problems / Pain Points",
            "2-3 additional pain points that reinforce the need",
            height=3
        )
        
        self.form_widgets["powerpoint_pitch"]["problem_stats"] = self._create_text_input(
            content,
            "Key Statistics & Evidence",
            "Compelling data points to include",
            height=2
        )
        
        # Section 2: Solution Attributes
        self._create_section_header(
            content,
            "🟢 Solution Attributes",
            "How does the solution specifically address the needs?"
        )
        
        self.form_widgets["powerpoint_pitch"]["solution_how"] = self._create_text_input(
            content,
            "How It Works",
            "Brief explanation of how the innovation solves the problem",
            height=3
        )
        
        self.form_widgets["powerpoint_pitch"]["solution_features"] = self._create_text_input(
            content,
            "Key Features & Capabilities",
            "3-5 key features that differentiate this solution",
            height=3
        )
        
        self.form_widgets["powerpoint_pitch"]["solution_benefits"] = self._create_text_input(
            content,
            "Quantified Benefits",
            "Specific benefits with numbers where possible",
            height=3
        )
        
        # Section 3: Value Proposition
        self._create_section_header(
            content,
            "💎 Value Proposition & Differentiation",
            "What is the core value message and competitive edge?"
        )
        
        self.form_widgets["powerpoint_pitch"]["value_prop"] = self._create_text_input(
            content,
            "Core Value Proposition",
            "One compelling sentence that captures the unique value",
            height=2
        )
        
        self.form_widgets["powerpoint_pitch"]["differentiators"] = self._create_text_input(
            content,
            "Key Differentiators vs. Competition",
            "What specifically makes this better than alternatives?",
            height=3
        )
        
        # Section 4: The Ask
        self._create_section_header(
            content,
            "🎯 The Ask",
            "What do you want from the audience?"
        )
        
        self.form_widgets["powerpoint_pitch"]["the_ask"] = self._create_text_input(
            content,
            "Investment/Resource Request",
            "What are you asking for?",
            height=2
        )
        
        self.form_widgets["powerpoint_pitch"]["use_of_funds"] = self._create_text_input(
            content,
            "Use of Funds / Resources",
            "How will the investment be used?",
            height=2
        )
        
        # Spacer
        tk.Frame(content, bg=self.bg_color, height=30).pack(fill=tk.X)
    
    # =========================================================================
    # PUBLIC METHODS
    # =========================================================================
    
    def set_output_type(self, output_type: str):
        """Switch to show the appropriate form for the output type."""
        self._show_form(output_type)
        self.current_output_type = output_type
    
    def _show_form(self, output_type: str):
        """Show the form for the specified output type, hide others."""
        for type_id, frame in self.form_frames.items():
            if type_id == output_type:
                frame.pack(fill=tk.BOTH, expand=True)
            else:
                frame.pack_forget()
    
    def get_values(self) -> Dict[str, str]:
        """Get all values from the current output type's form."""
        values = {}
        
        if self.current_output_type in self.form_widgets:
            widgets = self.form_widgets[self.current_output_type]
            for key, widget in widgets.items():
                if isinstance(widget, tk.Text):
                    values[key] = widget.get("1.0", tk.END).strip()
                elif isinstance(widget, tk.Entry):
                    values[key] = widget.get().strip()
        
        return values
    
    def get_all_values(self) -> Dict[str, Dict[str, str]]:
        """Get values from all output types (for persistence)."""
        all_values = {}
        
        for output_type, widgets in self.form_widgets.items():
            all_values[output_type] = {}
            for key, widget in widgets.items():
                if isinstance(widget, tk.Text):
                    all_values[output_type][key] = widget.get("1.0", tk.END).strip()
                elif isinstance(widget, tk.Entry):
                    all_values[output_type][key] = widget.get().strip()
        
        return all_values
    
    def clear(self):
        """Clear all form fields for all output types."""
        for output_type, widgets in self.form_widgets.items():
            for key, widget in widgets.items():
                if isinstance(widget, tk.Text):
                    widget.delete("1.0", tk.END)
                elif isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)
    
    def clear_current(self):
        """Clear only the current output type's form."""
        if self.current_output_type in self.form_widgets:
            widgets = self.form_widgets[self.current_output_type]
            for key, widget in widgets.items():
                if isinstance(widget, tk.Text):
                    widget.delete("1.0", tk.END)
                elif isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)


# =============================================================================
# PROMPT OUTPUT COMPONENT
# =============================================================================

class PromptOutputComponent(tk.Frame):
    """Component for displaying generated prompts with scrolling."""
    
    def __init__(self, parent, on_copy: Callable, on_export: Callable, on_generate: Callable = None):
        super().__init__(parent, bg=ADI_COLORS["white"])
        self.on_copy = on_copy
        self.on_export = on_export
        self.on_generate = on_generate
        self._create_widgets()
    
    def _create_widgets(self):
        header = tk.Frame(self, bg=ADI_COLORS["white"])
        header.pack(fill=tk.X)
        
        tk.Label(
            header,
            text="Generated Prompt",
            font=(UI_CONFIG["font_family"], 12, "bold"),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["dark_gray"]
        ).pack(side=tk.LEFT)
        
        tk.Button(
            header,
            text="📋 Copy to Clipboard",
            command=self.on_copy,
            bg=ADI_COLORS["success_green"],
            fg="black",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"], "bold"),
            padx=15,
            cursor="hand2"
        ).pack(side=tk.RIGHT)
        
         # Add Generate Prompt button here
        # Add Generate Prompt button if handler is provided
        if self.on_generate:
            tk.Button(
                header,
                text="🚀 Generate Prompt",
                command=self.on_generate,
                bg=ADI_COLORS["accent_orange"],
                fg="black",
                font=(UI_CONFIG["font_family"], 11, "bold"),
                padx=15,
                cursor="hand2"
            ).pack(side=tk.RIGHT)


        tk.Button(
            header,
            text="📤 Export",
            command=self.on_export,
            bg=ADI_COLORS["primary_blue"],
            fg="black",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            padx=10,
            cursor="hand2"
        ).pack(side=tk.RIGHT, padx=10)
        
        # Scrolled text area
        self.text_area = scrolledtext.ScrolledText(
            self,
            font=(UI_CONFIG["code_font_family"], UI_CONFIG["small_font_size"]),
            bg=ADI_COLORS["code_bg"],
            fg=ADI_COLORS["code_fg"],
            insertbackground=ADI_COLORS["accent_orange"],
            wrap=tk.WORD
        )
        self.text_area.pack(fill=tk.BOTH, expand=True, pady=10)
    
    def set_content(self, content):
        """Set the content of the text area with type safety."""
        self.text_area.delete("1.0", tk.END)
        # Defensive: ensure content is always a string
        if content is None:
            content = ""
        elif not isinstance(content, str):
            content = str(content)
        self.text_area.insert(tk.END, content)



    def get_content(self) -> str:
        return self.text_area.get("1.0", tk.END).strip()
    
    def clear(self):
        self.text_area.delete("1.0", tk.END)


# =============================================================================
# SCRIPT INPUT COMPONENT
# =============================================================================

class ScriptInputComponent(tk.Frame):
    """Component for script input and execution with scrolling."""
    
    def __init__(self, parent, on_execute: Callable, on_clear: Callable):
        super().__init__(parent, bg=ADI_COLORS["white"])
        self.on_execute = on_execute
        self.on_clear = on_clear
        self._create_widgets()
    
    def _create_widgets(self):
        info_frame = tk.Frame(self, bg=ADI_COLORS["light_blue"], pady=10)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(
            info_frame,
            text="💡 Paste the AI-generated Python script below. The script must create a workbook object named 'wb'. Do NOT include wb.save().",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            bg=ADI_COLORS["light_blue"],
            fg=ADI_COLORS["dark_blue"],
            wraplength=900
        ).pack(padx=10)
        
        tk.Label(
            self,
            text="Python Script Input:",
            font=(UI_CONFIG["font_family"], UI_CONFIG["body_font_size"], "bold"),
            bg=ADI_COLORS["white"],
            fg=ADI_COLORS["dark_gray"]
        ).pack(anchor="w", pady=(0, 5))
        
        # Scrolled text input
        self.script_input = scrolledtext.ScrolledText(
            self,
            font=(UI_CONFIG["code_font_family"], UI_CONFIG["body_font_size"]),
            bg=ADI_COLORS["code_bg"],
            fg=ADI_COLORS["code_fg"],
            insertbackground=ADI_COLORS["accent_orange"],
            wrap=tk.WORD
        )
        self.script_input.pack(fill=tk.BOTH, expand=True)
        
        btn_frame = tk.Frame(self, bg=ADI_COLORS["white"], pady=15)
        btn_frame.pack(fill=tk.X)
        
        tk.Button(
            btn_frame,
            text="🗑️ Clear",
            command=self.on_clear,
            bg=ADI_COLORS["medium_gray"],
            fg="black",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            width=10,
            height=2,
            cursor="hand2"
        ).pack(side=tk.LEFT)
        
        tk.Button(
            btn_frame,
            text="▶️ Execute & Save Excel",
            command=self.on_execute,
            bg=ADI_COLORS["accent_orange"],
            fg="black",
            font=(UI_CONFIG["font_family"], 13, "bold"),
            width=24,
            height=2,
            cursor="hand2"
        ).pack(side=tk.RIGHT)
    
    def get_script(self) -> str:
        return self.script_input.get("1.0", tk.END).strip()
    
    def clear(self):
        self.script_input.delete("1.0", tk.END)


# =============================================================================
# ERROR DIALOG COMPONENT
# =============================================================================

class ErrorDialogComponent(tk.Toplevel):
    """Error dialog for displaying script execution errors with scrolling."""
    
    def __init__(self, parent, error_message: str, on_export_log: Optional[Callable] = None):
        super().__init__(parent)
        self.on_export_log = on_export_log
        self.title("Script Execution Error")
        self.geometry("850x450")
        self.configure(bg=ADI_COLORS["white"])
        self._create_widgets(error_message)
    
    def _create_widgets(self, error_message: str):
        header = tk.Frame(self, bg=ADI_COLORS["error_red"], pady=10)
        header.pack(fill=tk.X)
        
        tk.Label(
            header,
            text="⚠️ Script Execution Error",
            font=(UI_CONFIG["font_family"], 14, "bold"),
            bg=ADI_COLORS["error_red"],
            fg="white"
        ).pack()
        
        tk.Label(
            header,
            text="Review the error below and fix your script",
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            bg=ADI_COLORS["error_red"],
            fg="white"
        ).pack()
        
        error_text = scrolledtext.ScrolledText(
            self,
            font=(UI_CONFIG["code_font_family"], UI_CONFIG["small_font_size"]),
            bg=ADI_COLORS["code_bg"],
            fg=ADI_COLORS["code_fg"]
        )
        error_text.insert(tk.END, error_message)
        error_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        btn_frame = tk.Frame(self, bg=ADI_COLORS["white"])
        btn_frame.pack(pady=10)
        
        if self.on_export_log:
            tk.Button(
                btn_frame,
                text="Export Log",
                command=self.on_export_log,
                bg=ADI_COLORS["primary_blue"],
                fg="black",
                font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
                padx=15,
                cursor="hand2"
            ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            btn_frame,
            text="Close",
            command=self.destroy,
            font=(UI_CONFIG["font_family"], UI_CONFIG["small_font_size"]),
            padx=15,
            cursor="hand2"
        ).pack(side=tk.LEFT)
