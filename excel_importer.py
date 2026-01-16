"""
excel_importer.py
Handles importing innovation data from Excel files.
Enhanced with output-specific field support for all output types.
"""

import os
from typing import Dict, Optional, Tuple, List
from dataclasses import dataclass
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


@dataclass
class ImportResult:
    """Result of Excel import operation."""
    success: bool
    data: Dict[str, str]
    specific_data: Dict[str, str]  # Output-specific fields
    errors: List[str]
    warnings: List[str]
    source_file: str
    output_type: str  # Detected output type


class ExcelImporter:
    """
    Imports innovation data from Excel files.
    Supports multiple Excel formats and field mappings including output-specific fields.
    """
    
    # =========================================================================
    # BASE FIELD MAPPINGS
    # =========================================================================
    
    FIELD_MAPPINGS = {
        # Primary fields
        "innovation name": "innovation_name",
        "name": "innovation_name",
        "project name": "innovation_name",
        
        "innovation description": "innovation_description",
        "description": "innovation_description",
        "project description": "innovation_description",
        
        "target industry": "industry",
        "industry": "industry",
        "market": "industry",
        "target market": "industry",
        "sector": "industry",
        
        "geographic scope": "geographic_scope",
        "geography": "geographic_scope",
        "region": "geographic_scope",
        "market region": "geographic_scope",
        
        "analysis timeframe": "analysis_timeframe",
        "analysis timeline": "analysis_timeframe",
        "timeframe": "analysis_timeframe",
        "timeline": "analysis_timeframe",
        "projection period": "analysis_timeframe",
        
        "innovation stage": "innovation_stage",
        "stage": "innovation_stage",
        "development stage": "innovation_stage",
        "project stage": "innovation_stage",
        "trl": "innovation_stage",
        
        "currency": "currency",
        "output type": "output_type",
    }
    
    # =========================================================================
    # OUTPUT-SPECIFIC FIELD MAPPINGS
    # =========================================================================
    
    SPECIFIC_FIELD_MAPPINGS = {
        # ----- Excel Value Model Fields -----
        "primary value drivers": "value_drivers",
        "value drivers": "value_drivers",
        "value calculation factors": "value_factors",
        "calculation factors": "value_factors",
        "stakeholder identification": "stakeholders",
        "key stakeholders": "stakeholders",
        "stakeholders": "stakeholders",
        "value allocation guidance": "value_allocation",
        "value allocation": "value_allocation",
        "segments to include": "segments_include",
        "included segments": "segments_include",
        "segments to exclude": "segments_exclude",
        "excluded segments": "segments_exclude",
        "growth rate guidance": "growth_assumptions",
        "growth assumptions": "growth_assumptions",
        
        # ----- PowerPoint Pitch Fields -----
        "primary problem statement": "problem_primary",
        "problem statement": "problem_primary",
        "supporting problems": "problem_secondary",
        "pain points": "problem_secondary",
        "supporting problems / pain points": "problem_secondary",
        "key statistics": "problem_stats",
        "key statistics & evidence": "problem_stats",
        "evidence": "problem_stats",
        "how it works": "solution_how",
        "solution description": "solution_how",
        "key features": "solution_features",
        "key features & capabilities": "solution_features",
        "features": "solution_features",
        "quantified benefits": "solution_benefits",
        "benefits": "solution_benefits",
        "core value proposition": "value_prop",
        "value proposition": "value_prop",
        "key differentiators": "differentiators",
        "differentiators": "differentiators",
        "differentiation": "differentiators",
        "investment request": "the_ask",
        "the ask": "the_ask",
        "resource request": "the_ask",
        "investment/resource request": "the_ask",
        "use of funds": "use_of_funds",
        "use of funds / resources": "use_of_funds",
        
        # ----- Word Deep Dive Fields -----
        "executive summary guidance": "exec_summary",
        "executive summary": "exec_summary",
        "brief overview & decision statement": "exec_summary",
        "purpose & scope": "purpose_scope",
        "purpose and scope": "purpose_scope",
        "scope": "purpose_scope",
        "objective": "purpose_scope",
        "technology & market overview": "tech_market",
        "technology and market": "tech_market",
        "tech market": "tech_market",
        "current state of technology & market": "tech_market",
        "competitive landscape": "competitive",
        "competition": "competitive",
        "competitor summary": "competitive",
        "rationale for go decision": "rationale",
        "go rationale": "rationale",
        "rationale": "rationale",
        "key assessment findings": "rationale",
        "risks & unknowns": "risks",
        "risks and unknowns": "risks",
        "key risks": "risks",
        "technology triggers": "triggers",
        "triggers": "triggers",
        "technology triggers & market dynamics": "triggers",
        "collaboration opportunities": "collaboration",
        "collaboration": "collaboration",
        "partners": "collaboration",
        "potential partners": "collaboration",
        "next steps": "next_steps",
        "follow-up": "next_steps",
        "immediate actions": "next_steps",
    }
    
    # Required fields for validation
    REQUIRED_FIELDS = ["innovation_name", "innovation_description", "industry"]
    
    def __init__(self):
        self.last_import: Optional[ImportResult] = None
    
    def import_from_excel(self, filepath: str) -> ImportResult:
        """
        Import innovation data from an Excel file.
        
        Supports two formats:
        1. Two-column format: Column A = field names, Column B = values
        2. Header row format: First row = headers, subsequent rows = data
        
        Args:
            filepath: Path to the Excel file
        
        Returns:
            ImportResult with extracted data including output-specific fields
        """
        errors = []
        warnings = []
        data = {}
        specific_data = {}
        output_type = "excel_value_model"  # Default
        
        if not os.path.exists(filepath):
            return ImportResult(
                success=False,
                data={},
                specific_data={},
                errors=[f"File not found: {filepath}"],
                warnings=[],
                source_file=filepath,
                output_type=output_type
            )
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Detect format and extract data
            data, specific_data, output_type = self._extract_data(ws, warnings)
            
            # Validate required fields
            missing_fields = []
            for field in self.REQUIRED_FIELDS:
                if field not in data or not data[field]:
                    missing_fields.append(field)
            
            if missing_fields:
                warnings.append(f"Missing recommended fields: {', '.join(missing_fields)}")
            
            wb.close()
            
            self.last_import = ImportResult(
                success=True,
                data=data,
                specific_data=specific_data,
                errors=errors,
                warnings=warnings,
                source_file=filepath,
                output_type=output_type
            )
            return self.last_import
            
        except Exception as e:
            return ImportResult(
                success=False,
                data={},
                specific_data={},
                errors=[f"Error reading Excel file: {str(e)}"],
                warnings=[],
                source_file=filepath,
                output_type=output_type
            )
    
    def _extract_data(self, ws, warnings: List[str]) -> Tuple[Dict[str, str], Dict[str, str], str]:
        """
        Extract data from worksheet, auto-detecting format.
        
        Returns:
            Tuple of (base_data, specific_data, output_type)
        """
        data = {}
        specific_data = {}
        output_type = "excel_value_model"
        
        # Try two-column format first (Column A = labels, Column B = values)
        for row in ws.iter_rows(min_row=1, max_col=2):
            if len(row) >= 2:
                label_cell = row[0]
                value_cell = row[1]
                
                if label_cell.value and value_cell.value:
                    label = str(label_cell.value).strip().lower()
                    value = str(value_cell.value).strip()
                    
                    # Skip section headers
                    if label.startswith("---") or label.startswith("==="):
                        continue
                    
                    # Skip metadata labels
                    if label in ["field", "value", "notes", "export metadata"]:
                        continue
                    
                    # Check base field mappings
                    field_name = self.FIELD_MAPPINGS.get(label)
                    if field_name:
                        if field_name == "output_type":
                            output_type = value
                        else:
                            data[field_name] = value
                        continue
                    
                    # Check output-specific field mappings
                    specific_field = self.SPECIFIC_FIELD_MAPPINGS.get(label)
                    if specific_field:
                        specific_data[specific_field] = value
                        continue
                    
                    # Check for "(Other Detail)" patterns
                    if "(other detail)" in label or "(other)" in label:
                        base_field = label.replace("(other detail)", "").replace("(other)", "").strip()
                        base_field = base_field.replace(" ", "_")
                        # Store but don't warn - these are valid "Other" details
                        continue
                    
                    # Unknown field - try to categorize
                    if label and not label.startswith("unnamed") and not label.startswith("export"):
                        clean_label = label.replace(" ", "_")
                        
                        # Try to categorize as specific field based on keywords
                        specific_keywords = [
                            "value", "stakeholder", "segment", "growth", 
                            "problem", "solution", "feature", "benefit",
                            "ask", "fund", "summary", "scope", "risk",
                            "trigger", "collaboration", "step", "competitor",
                            "rationale", "market", "technology"
                        ]
                        
                        if any(kw in label for kw in specific_keywords):
                            specific_data[f"custom_{clean_label}"] = value
                        else:
                            data[f"custom_{clean_label}"] = value
                        
                        warnings.append(f"Unknown field imported as custom: {label}")
        
        return data, specific_data, output_type
    
    def get_template_path(self) -> str:
        """Get path to the Excel template file."""
        import sys
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, "Input_template.xlsx")
    
    def create_template(self, filepath: str, output_type: str = "excel_value_model") -> bool:
        """Create an Excel template appropriate for the selected output type."""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Innovation Input"
            
            # Style for headers
            header_fill = PatternFill(start_color="0067B9", end_color="0067B9", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            section_fill = PatternFill(start_color="4A9BD9", end_color="4A9BD9", fill_type="solid")
            section_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            ws["A1"] = "Field"
            ws["B1"] = "Value"
            ws["C1"] = "Notes"
            for col in ["A", "B", "C"]:
                ws[f"{col}1"].fill = header_fill
                ws[f"{col}1"].font = header_font
            
            # Get template fields based on output type
            template_fields = self._get_template_fields_for_type(output_type)
            
            # Write template fields
            for i, (field, default, note) in enumerate(template_fields, start=2):
                # Check if this is a section header
                if field.startswith("---"):
                    ws.cell(row=i, column=1, value=field)
                    ws.cell(row=i, column=1).fill = section_fill
                    ws.cell(row=i, column=1).font = section_font
                    ws.merge_cells(f'A{i}:C{i}')
                elif field == "":
                    # Empty row for spacing
                    pass
                else:
                    ws.cell(row=i, column=1, value=field)
                    ws.cell(row=i, column=2, value=default)
                    ws.cell(row=i, column=3, value=note)
                    ws.cell(row=i, column=3).font = Font(color="666666", italic=True)
            
            # Set column widths
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 55
            
            wb.save(filepath)
            return True
        
        except Exception as e:
            print(f"Error creating template: {e}")
            return False

    def _get_template_fields_for_type(self, output_type: str) -> list:
        """Get template fields specific to the output type."""
        
        # Common base fields for all types
        base_fields = [
            ("--- REQUIRED FIELDS ---", "", ""),
            ("Innovation Name", "", "Required - Name of your innovation"),
            ("Innovation Description", "", "Required - 2-4 sentences describing the innovation"),
            ("Target Industry", "", "Required - e.g., Healthcare / Medical Devices"),
            ("", "", ""),
            ("--- ANALYSIS PARAMETERS ---", "", ""),
            ("Geographic Scope", "Global", "Options: United States, North America, Europe, Asia-Pacific, Global, Other"),
            ("Analysis Timeframe", "Year 1 at Scale", "Options: Year 1 at Scale, 3-Year Projection, 5-Year Projection, 10-Year Projection, Other"),
            ("Innovation Stage", "Concept", "Options: Concept, Prototype, Pilot, Commercial, Growth, Other"),
            ("Currency", "USD", "Options: USD, EUR, GBP, JPY, Other"),
            ("Output Type", output_type, "excel_value_model, powerpoint_pitch, or word_gonogo"),
            ("", "", ""),
        ]
        
        if output_type == "excel_value_model":
            specific_fields = [
                ("--- VALUE MODEL INPUTS ---", "", ""),
                ("Primary Value Drivers", "", "List 3-5 main ways this innovation creates value (e.g., cost reduction, revenue increase)"),
                ("Value Calculation Factors", "", "What factors/metrics should be used to quantify each driver?"),
                ("", "", ""),
                ("Stakeholder Identification", "", "List key stakeholders and how each captures value"),
                ("Value Allocation Guidance", "", "How should value be allocated among stakeholders?"),
                ("", "", ""),
                ("Segments to Include", "", "List specific market segments to analyze"),
                ("Segments to Exclude", "", "List any segments to exclude and rationale"),
                ("", "", ""),
                ("Growth Rate Guidance", "", "Expected growth drivers, market expansion factors, or constraints"),
            ]
        
        elif output_type == "powerpoint_pitch":
            specific_fields = [
                ("--- PROBLEM STATEMENTS ---", "", ""),
                ("Primary Problem Statement", "", "The #1 problem this solves - make it compelling and specific"),
                ("Supporting Problems / Pain Points", "", "2-3 additional pain points that reinforce the need"),
                ("Key Statistics & Evidence", "", "Compelling data points about the problem (include sources)"),
                ("", "", ""),
                ("--- SOLUTION ATTRIBUTES ---", "", ""),
                ("How It Works", "", "Brief explanation of how the innovation solves the problem"),
                ("Key Features & Capabilities", "", "3-5 key features that differentiate this solution"),
                ("Quantified Benefits", "", "Specific benefits with numbers where possible (e.g., 40% reduction)"),
                ("", "", ""),
                ("--- VALUE PROPOSITION ---", "", ""),
                ("Core Value Proposition", "", "One compelling sentence that captures the unique value"),
                ("Key Differentiators", "", "What specifically makes this better than alternatives?"),
                ("", "", ""),
                ("--- THE ASK ---", "", ""),
                ("Investment/Resource Request", "", "What are you asking for? (e.g., $2M seed funding)"),
                ("Use of Funds / Resources", "", "How will the investment be used? (e.g., 40% R&D, 30% clinical trials)"),
            ]
        
        elif output_type == "word_gonogo":
            specific_fields = [
                ("--- DEEP DIVE SECTION GUIDANCE ---", "", ""),
                ("Executive Summary Guidance", "", "Key finding and recommendation summary"),
                ("Purpose & Scope", "", "Objective and exploration focus - what questions will be addressed"),
                ("", "", ""),
                ("Technology & Market Overview", "", "Key technologies, maturity levels, market size, customer segments"),
                ("Competitive Landscape", "", "Key competitors, their offerings, positioning"),
                ("", "", ""),
                ("Rationale for Go Decision", "", "Most important findings supporting the recommendation to proceed"),
                ("Risks & Unknowns", "", "Major risks, IP considerations, mitigation strategies"),
                ("", "", ""),
                ("Technology Triggers", "", "Breakthroughs or standards that would change the opportunity"),
                ("Collaboration Opportunities", "", "Potential partners, companies to watch, consortium opportunities"),
                ("", "", ""),
                ("Next Steps", "", "Immediate actions to launch Exploration phase"),
            ]
        
        else:
            specific_fields = []
        
        return base_fields + specific_fields


# Global instance
excel_importer = ExcelImporter()
