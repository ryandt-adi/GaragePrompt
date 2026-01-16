"""
excel_processor.py
Comprehensive Excel file operations and script execution for Analog Garage Workbench.
Includes all chart types, styling utilities, and business modeling helpers.
"""

import traceback
from typing import Dict, Any, Optional, Tuple, List, Union
from enum import Enum

# =============================================================================
# OPENPYXL IMPORTS - SAFE VERSION
# =============================================================================

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Styles
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, GradientFill, Color, Protection
)

# Number format constants - define our own to avoid version issues
NUMBER_FORMATS = {
    "GENERAL": "General",
    "NUMBER": "0",
    "NUMBER_00": "0.00",
    "NUMBER_COMMA": "#,##0",
    "NUMBER_COMMA_00": "#,##0.00",
    "PERCENTAGE": "0%",
    "PERCENTAGE_00": "0.00%",
    "CURRENCY_USD": '"$"#,##0.00',
    "CURRENCY_USD_SIMPLE": "$#,##0",
    "CURRENCY_EUR": '[$€-407]#,##0.00',
    "CURRENCY_GBP": '[$£-809]#,##0.00',
    "DATE_YYYYMMDD": "YYYY-MM-DD",
    "DATE_DDMMYY": "DD/MM/YY",
    "DATE_MMDDYYYY": "MM/DD/YYYY",
    "DATE_MMMDDYYYY": "MMM DD, YYYY",
    "DATETIME": "YYYY-MM-DD HH:MM:SS",
    "TIME": "HH:MM:SS",
    "TIME_12H": "h:mm:ss AM/PM",
    "SCIENTIFIC": "0.00E+00",
    "TEXT": "@",
    "ACCOUNTING": '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
}

# Utilities
from openpyxl.utils import get_column_letter, column_index_from_string

# Charts - ALL TYPES
from openpyxl.chart import (
    # Basic Charts
    BarChart, BarChart3D,
    LineChart, LineChart3D,
    AreaChart, AreaChart3D,
    PieChart, PieChart3D,
    DoughnutChart,
    ScatterChart,
    BubbleChart,
    RadarChart,
    StockChart,
    SurfaceChart, SurfaceChart3D,
    # Chart Components
    Reference,
    Series,
)

# Chart Labels
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend

# Conditional Formatting
from openpyxl.formatting.rule import (
    ColorScaleRule, FormulaRule, CellIsRule,
    IconSetRule, DataBarRule, Rule
)
from openpyxl.styles.differential import DifferentialStyle

# Data Validation
from openpyxl.worksheet.datavalidation import DataValidation

# Tables
from openpyxl.worksheet.table import Table, TableStyleInfo

# Comments
from openpyxl.comments import Comment

# Images - optional
try:
    from openpyxl.drawing.image import Image as ExcelImage
    IMAGES_SUPPORTED = True
except ImportError:
    IMAGES_SUPPORTED = False

# Date utilities
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta


# =============================================================================
# ENUMS AND CONSTANTS
# =============================================================================

class ChartType(Enum):
    """All supported chart types in openpyxl."""
    BAR = "bar"
    BAR_3D = "bar3d"
    LINE = "line"
    LINE_3D = "line3d"
    AREA = "area"
    AREA_3D = "area3d"
    PIE = "pie"
    PIE_3D = "pie3d"
    DOUGHNUT = "doughnut"
    SCATTER = "scatter"
    BUBBLE = "bubble"
    RADAR = "radar"
    STOCK = "stock"
    SURFACE = "surface"
    SURFACE_3D = "surface3d"


class NumberFormat(Enum):
    """Common number formats for business modeling."""
    GENERAL = "General"
    NUMBER = "#,##0"
    NUMBER_2DP = "#,##0.00"
    CURRENCY = "$#,##0"
    CURRENCY_2DP = "$#,##0.00"
    CURRENCY_NEGATIVE_RED = "$#,##0_);[Red]($#,##0)"
    ACCOUNTING = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    PERCENTAGE = "0%"
    PERCENTAGE_2DP = "0.00%"
    SCIENTIFIC = "0.00E+00"
    DATE_SHORT = "MM/DD/YYYY"
    DATE_LONG = "MMMM D, YYYY"
    DATE_ISO = "YYYY-MM-DD"
    TIME = "HH:MM:SS"
    DATETIME = "MM/DD/YYYY HH:MM"
    TEXT = "@"
    THOUSANDS = '#,##0,"K"'
    MILLIONS = '#,##0,,"M"'
    BILLIONS = '#,##0,,,"B"'


# Brand colors
BRAND_COLORS = {
    "primary_blue": "0067B9",
    "dark_blue": "003D6A",
    "light_blue": "4A9BD9",
    "accent_orange": "FF6600",
    "success_green": "28A745",
    "error_red": "DC3545",
    "warning_yellow": "FFC107",
    "white": "FFFFFF",
    "light_gray": "F5F5F5",
    "dark_gray": "333333",
}


# =============================================================================
# DATA CLASSES (using simple classes for compatibility)
# =============================================================================

class ExecutionResult:
    """Result of script execution."""
    
    def __init__(self, success: bool, workbook: Optional[Workbook], 
                 error_message: str, traceback_str: str):
        self.success = success
        self.workbook = workbook
        self.error_message = error_message
        self.traceback = traceback_str
    
    @property
    def has_error(self) -> bool:
        return not self.success


class ChartConfig:
    """Configuration for chart creation."""
    
    def __init__(
        self,
        chart_type: ChartType,
        title: str,
        data_range: str,
        categories_range: Optional[str] = None,
        position: str = "E1",
        width: float = 15,
        height: float = 10,
        style: int = 10,
        legend_position: str = "r"
    ):
        self.chart_type = chart_type
        self.title = title
        self.data_range = data_range
        self.categories_range = categories_range
        self.position = position
        self.width = width
        self.height = height
        self.style = style
        self.legend_position = legend_position


class TableConfig:
    """Configuration for Excel table creation."""
    
    def __init__(
        self,
        name: str,
        data_range: str,
        style: str = "TableStyleMedium2",
        show_first_col: bool = False,
        show_last_col: bool = False,
        show_row_stripes: bool = True,
        show_col_stripes: bool = False
    ):
        self.name = name
        self.data_range = data_range
        self.style = style
        self.show_first_col = show_first_col
        self.show_last_col = show_last_col
        self.show_row_stripes = show_row_stripes
        self.show_col_stripes = show_col_stripes


# =============================================================================
# STYLE FACTORY
# =============================================================================

class StyleFactory:
    """Factory for creating common Excel styles."""
    
    @staticmethod
    def header_style(
        bg_color: str = BRAND_COLORS["primary_blue"],
        font_color: str = BRAND_COLORS["white"],
        bold: bool = True,
        font_size: int = 11
    ) -> dict:
        """Create header cell style."""
        return {
            "font": Font(bold=bold, color=font_color, size=font_size),
            "fill": PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": StyleFactory.thin_border()
        }
    
    @staticmethod
    def subheader_style(
        bg_color: str = BRAND_COLORS["light_blue"],
        font_color: str = BRAND_COLORS["dark_blue"]
    ) -> dict:
        """Create subheader cell style."""
        return {
            "font": Font(bold=True, color=font_color, size=10),
            "fill": PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": StyleFactory.thin_border()
        }
    
    @staticmethod
    def data_style(align: str = "left") -> dict:
        """Create standard data cell style."""
        return {
            "font": Font(size=10),
            "alignment": Alignment(horizontal=align, vertical="center"),
            "border": StyleFactory.thin_border()
        }
    
    @staticmethod
    def currency_style() -> dict:
        """Create currency cell style."""
        return {
            "font": Font(size=10),
            "alignment": Alignment(horizontal="right", vertical="center"),
            "border": StyleFactory.thin_border(),
            "number_format": NumberFormat.CURRENCY_2DP.value
        }
    
    @staticmethod
    def percentage_style() -> dict:
        """Create percentage cell style."""
        return {
            "font": Font(size=10),
            "alignment": Alignment(horizontal="right", vertical="center"),
            "border": StyleFactory.thin_border(),
            "number_format": NumberFormat.PERCENTAGE_2DP.value
        }
    
    @staticmethod
    def highlight_positive() -> dict:
        """Style for positive values."""
        return {
            "font": Font(color=BRAND_COLORS["success_green"], bold=True),
            "fill": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        }
    
    @staticmethod
    def highlight_negative() -> dict:
        """Style for negative values."""
        return {
            "font": Font(color=BRAND_COLORS["error_red"], bold=True),
            "fill": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        }
    
    @staticmethod
    def thin_border() -> Border:
        """Create thin border."""
        thin = Side(style="thin", color="000000")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    
    @staticmethod
    def thick_border() -> Border:
        """Create thick border."""
        thick = Side(style="medium", color="000000")
        return Border(left=thick, right=thick, top=thick, bottom=thick)
    
    @staticmethod
    def no_border() -> Border:
        """Create no border."""
        return Border()
    
    @staticmethod
    def title_style() -> dict:
        """Create title cell style."""
        return {
            "font": Font(bold=True, size=16, color=BRAND_COLORS["dark_blue"]),
            "alignment": Alignment(horizontal="left", vertical="center")
        }
    
    @staticmethod
    def kpi_value_style(color: str = BRAND_COLORS["primary_blue"]) -> dict:
        """Style for KPI values."""
        return {
            "font": Font(bold=True, size=24, color=color),
            "alignment": Alignment(horizontal="center", vertical="center")
        }
    
    @staticmethod
    def kpi_label_style() -> dict:
        """Style for KPI labels."""
        return {
            "font": Font(size=10, color=BRAND_COLORS["dark_gray"]),
            "alignment": Alignment(horizontal="center", vertical="center")
        }
    
    @staticmethod
    def apply_style(cell, style_dict: dict):
        """Apply a style dictionary to a cell."""
        for key, value in style_dict.items():
            setattr(cell, key, value)


# =============================================================================
# CHART FACTORY
# =============================================================================

class ChartFactory:
    """Factory for creating various chart types."""
    
    @staticmethod
    def create_chart(
        chart_type: ChartType,
        title: str = "",
        style: int = 10
    ):
        """Create a chart object of the specified type."""
        chart_map = {
            ChartType.BAR: BarChart,
            ChartType.BAR_3D: BarChart3D,
            ChartType.LINE: LineChart,
            ChartType.LINE_3D: LineChart3D,
            ChartType.AREA: AreaChart,
            ChartType.AREA_3D: AreaChart3D,
            ChartType.PIE: PieChart,
            ChartType.PIE_3D: PieChart3D,
            ChartType.DOUGHNUT: DoughnutChart,
            ChartType.SCATTER: ScatterChart,
            ChartType.BUBBLE: BubbleChart,
            ChartType.RADAR: RadarChart,
            ChartType.STOCK: StockChart,
            ChartType.SURFACE: SurfaceChart,
            ChartType.SURFACE_3D: SurfaceChart3D,
        }
        
        chart_class = chart_map.get(chart_type)
        if not chart_class:
            raise ValueError(f"Unknown chart type: {chart_type}")
        
        chart = chart_class()
        chart.title = title
        chart.style = style
        return chart
    
    @staticmethod
    def create_bar_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: Optional[str] = None,
        title: str = "",
        position: str = "E1",
        bar_type: str = "col",
        stacked: bool = False,
        width: float = 15,
        height: float = 10
    ) -> BarChart:
        """Create a bar/column chart."""
        chart = BarChart()
        chart.type = bar_type
        chart.grouping = "stacked" if stacked else "clustered"
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_line_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: Optional[str] = None,
        title: str = "",
        position: str = "E1",
        smooth: bool = False,
        width: float = 15,
        height: float = 10
    ) -> LineChart:
        """Create a line chart."""
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)
        
        for series in chart.series:
            series.smooth = smooth
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_pie_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: str,
        title: str = "",
        position: str = "E1",
        width: float = 12,
        height: float = 10,
        show_percent: bool = True
    ) -> PieChart:
        """Create a pie chart."""
        chart = PieChart()
        chart.title = title
        chart.width = width
        chart.height = height
        
        data = Reference(ws, range_string=data_range)
        cats = Reference(ws, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if show_percent:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_doughnut_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: str,
        title: str = "",
        position: str = "E1",
        hole_size: int = 50,
        width: float = 12,
        height: float = 10
    ) -> DoughnutChart:
        """Create a doughnut chart."""
        chart = DoughnutChart()
        chart.title = title
        chart.width = width
        chart.height = height
        chart.holeSize = hole_size
        
        data = Reference(ws, range_string=data_range)
        cats = Reference(ws, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_scatter_chart(
        ws: Worksheet,
        x_range: str,
        y_range: str,
        title: str = "",
        position: str = "E1",
        x_title: str = "",
        y_title: str = "",
        width: float = 15,
        height: float = 10
    ) -> ScatterChart:
        """Create a scatter chart."""
        chart = ScatterChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        
        xvalues = Reference(ws, range_string=x_range)
        yvalues = Reference(ws, range_string=y_range)
        
        series = Series(yvalues, xvalues, title="Data")
        chart.series.append(series)
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_area_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: Optional[str] = None,
        title: str = "",
        position: str = "E1",
        stacked: bool = False,
        width: float = 15,
        height: float = 10
    ) -> AreaChart:
        """Create an area chart."""
        chart = AreaChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        chart.grouping = "stacked" if stacked else "standard"
        
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)
        
        ws.add_chart(chart, position)
        return chart
    
    @staticmethod
    def create_radar_chart(
        ws: Worksheet,
        data_range: str,
        categories_range: str,
        title: str = "",
        position: str = "E1",
        radar_type: str = "standard",
        width: float = 12,
        height: float = 10
    ) -> RadarChart:
        """Create a radar chart."""
        chart = RadarChart()
        chart.title = title
        chart.type = radar_type
        chart.width = width
        chart.height = height
        
        data = Reference(ws, range_string=data_range)
        cats = Reference(ws, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, position)
        return chart


# =============================================================================
# FORMULA HELPERS FOR BUSINESS MODELING
# =============================================================================

class FormulaHelper:
    """Helper class for Excel formulas commonly used in business modeling."""
    
    # Aggregation
    @staticmethod
    def sum(range_str: str) -> str:
        return f"=SUM({range_str})"
    
    @staticmethod
    def average(range_str: str) -> str:
        return f"=AVERAGE({range_str})"
    
    @staticmethod
    def count(range_str: str) -> str:
        return f"=COUNT({range_str})"
    
    @staticmethod
    def counta(range_str: str) -> str:
        return f"=COUNTA({range_str})"
    
    @staticmethod
    def min_val(range_str: str) -> str:
        return f"=MIN({range_str})"
    
    @staticmethod
    def max_val(range_str: str) -> str:
        return f"=MAX({range_str})"
    
    # Conditional
    @staticmethod
    def countif(range_str: str, criteria: str) -> str:
        return f'=COUNTIF({range_str},"{criteria}")'
    
    @staticmethod
    def sumif(range_str: str, criteria: str, sum_range: str) -> str:
        return f'=SUMIF({range_str},"{criteria}",{sum_range})'
    
    @staticmethod
    def averageif(range_str: str, criteria: str, avg_range: str) -> str:
        return f'=AVERAGEIF({range_str},"{criteria}",{avg_range})'
    
    # Lookup
    @staticmethod
    def vlookup(lookup_value: str, table_range: str, col_index: int, exact: bool = False) -> str:
        match_type = "FALSE" if exact else "TRUE"
        return f"=VLOOKUP({lookup_value},{table_range},{col_index},{match_type})"
    
    @staticmethod
    def hlookup(lookup_value: str, table_range: str, row_index: int, exact: bool = False) -> str:
        match_type = "FALSE" if exact else "TRUE"
        return f"=HLOOKUP({lookup_value},{table_range},{row_index},{match_type})"
    
    @staticmethod
    def index_match(lookup_value: str, lookup_range: str, return_range: str) -> str:
        return f"=INDEX({return_range},MATCH({lookup_value},{lookup_range},0))"
    
    # Logic
    @staticmethod
    def if_formula(condition: str, true_value: str, false_value: str) -> str:
        return f"=IF({condition},{true_value},{false_value})"
    
    @staticmethod
    def iferror(formula: str, error_value: str = '""') -> str:
        # Remove leading = if present
        formula = formula.lstrip('=')
        return f"=IFERROR({formula},{error_value})"
    
    @staticmethod
    def and_formula(*conditions) -> str:
        return f"=AND({','.join(conditions)})"
    
    @staticmethod
    def or_formula(*conditions) -> str:
        return f"=OR({','.join(conditions)})"
    
    # Financial
    @staticmethod
    def npv(rate: str, values_range: str) -> str:
        return f"=NPV({rate},{values_range})"
    
    @staticmethod
    def irr(values_range: str, guess: str = "0.1") -> str:
        return f"=IRR({values_range},{guess})"
    
    @staticmethod
    def xirr(values_range: str, dates_range: str, guess: str = "0.1") -> str:
        return f"=XIRR({values_range},{dates_range},{guess})"
    
    @staticmethod
    def xnpv(rate: str, values_range: str, dates_range: str) -> str:
        return f"=XNPV({rate},{values_range},{dates_range})"
    
    @staticmethod
    def pmt(rate: str, nper: str, pv: str, fv: str = "0", type_: str = "0") -> str:
        return f"=PMT({rate},{nper},{pv},{fv},{type_})"
    
    @staticmethod
    def pv(rate: str, nper: str, pmt: str, fv: str = "0") -> str:
        return f"=PV({rate},{nper},{pmt},{fv})"
    
    @staticmethod
    def fv(rate: str, nper: str, pmt: str, pv: str = "0") -> str:
        return f"=FV({rate},{nper},{pmt},{pv})"
    
    @staticmethod
    def cagr(start_value: str, end_value: str, periods: str) -> str:
        return f"=(({end_value}/{start_value})^(1/{periods}))-1"
    
    @staticmethod
    def payback_period(initial_investment: str, annual_cash_flow: str) -> str:
        return f"={initial_investment}/{annual_cash_flow}"
    
    # Statistical
    @staticmethod
    def stdev(range_str: str) -> str:
        return f"=STDEV({range_str})"
    
    @staticmethod
    def var(range_str: str) -> str:
        return f"=VAR({range_str})"
    
    @staticmethod
    def percentile(range_str: str, k: float) -> str:
        return f"=PERCENTILE({range_str},{k})"
    
    @staticmethod
    def correl(range1: str, range2: str) -> str:
        return f"=CORREL({range1},{range2})"
    
    @staticmethod
    def median(range_str: str) -> str:
        return f"=MEDIAN({range_str})"
    
    # Growth
    @staticmethod
    def growth_rate(new_value: str, old_value: str) -> str:
        return f"=({new_value}-{old_value})/{old_value}"
    
    @staticmethod
    def year_over_year(current: str, prior: str) -> str:
        return f"=({current}-{prior})/{prior}"
    
    # Rounding
    @staticmethod
    def round_val(value: str, decimals: int = 0) -> str:
        return f"=ROUND({value},{decimals})"
    
    @staticmethod
    def roundup(value: str, decimals: int = 0) -> str:
        return f"=ROUNDUP({value},{decimals})"
    
    @staticmethod
    def rounddown(value: str, decimals: int = 0) -> str:
        return f"=ROUNDDOWN({value},{decimals})"


# =============================================================================
# BUSINESS MODELING UTILITIES
# =============================================================================

class BusinessModelBuilder:
    """Utilities for building business model components in Excel."""
    
    @staticmethod
    def create_kpi_card(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        title: str,
        value: Union[str, float],
        subtitle: str = "",
        color: str = BRAND_COLORS["primary_blue"]
    ):
        """Create a KPI card in the worksheet."""
        value_cell = ws.cell(row=start_row, column=start_col, value=value)
        value_cell.font = Font(bold=True, size=24, color=color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        title_cell = ws.cell(row=start_row + 1, column=start_col, value=title)
        title_cell.font = Font(size=10, color=BRAND_COLORS["dark_gray"])
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if subtitle:
            sub_cell = ws.cell(row=start_row + 2, column=start_col, value=subtitle)
            sub_cell.font = Font(size=8, color=BRAND_COLORS["dark_gray"], italic=True)
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def create_data_table(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        headers: List[str],
        data: List[List[Any]],
        table_name: Optional[str] = None,
        style: str = "TableStyleMedium2"
    ) -> int:
        """Create a formatted data table. Returns the ending row."""
        # Write headers
        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + i, value=header)
            StyleFactory.apply_style(cell, StyleFactory.header_style())
        
        # Write data
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=value)
                StyleFactory.apply_style(cell, StyleFactory.data_style())
        
        end_row = start_row + len(data)
        end_col = start_col + len(headers) - 1
        
        # Create Excel table if name provided
        if table_name:
            table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            table = Table(displayName=table_name, ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name=style, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(table)
        
        return end_row
    
    @staticmethod
    def create_scenario_table(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        scenarios: Dict[str, Dict[str, float]],
        metrics: List[str]
    ):
        """Create a scenario comparison table."""
        scenario_names = list(scenarios.keys())
        
        # Header row
        ws.cell(row=start_row, column=start_col, value="Metric")
        StyleFactory.apply_style(
            ws.cell(row=start_row, column=start_col),
            StyleFactory.header_style()
        )
        
        for i, scenario in enumerate(scenario_names):
            cell = ws.cell(row=start_row, column=start_col + 1 + i, value=scenario)
            StyleFactory.apply_style(cell, StyleFactory.header_style())
        
        # Metric rows
        for i, metric in enumerate(metrics):
            ws.cell(row=start_row + 1 + i, column=start_col, value=metric)
            for j, scenario in enumerate(scenario_names):
                value = scenarios[scenario].get(metric, "")
                ws.cell(row=start_row + 1 + i, column=start_col + 1 + j, value=value)
    
    @staticmethod
    def create_assumption_log(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        assumptions: List[Dict[str, str]]
    ):
        """Create an assumption log table."""
        headers = ["ID", "Assumption", "Value", "Source", "Confidence", "Last Updated"]
        
        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + i, value=header)
            StyleFactory.apply_style(cell, StyleFactory.header_style())
        
        for i, assumption in enumerate(assumptions):
            row = start_row + 1 + i
            ws.cell(row=row, column=start_col, value=f"A{i+1}")
            ws.cell(row=row, column=start_col + 1, value=assumption.get("name", ""))
            ws.cell(row=row, column=start_col + 2, value=assumption.get("value", ""))
            ws.cell(row=row, column=start_col + 3, value=assumption.get("source", ""))
            ws.cell(row=row, column=start_col + 4, value=assumption.get("confidence", "Medium"))
            ws.cell(row=row, column=start_col + 5, value=assumption.get("updated", ""))
    
    @staticmethod
    def create_waterfall_data(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        items: List[Tuple[str, float]],
        title: str = "Bridge Analysis"
    ):
        """Create data structure for waterfall chart."""
        ws.cell(row=start_row, column=start_col, value=title)
        StyleFactory.apply_style(
            ws.cell(row=start_row, column=start_col),
            StyleFactory.title_style()
        )
        
        headers = ["Item", "Value", "Start", "End", "Increase", "Decrease"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row + 1, column=start_col + i, value=header)
            StyleFactory.apply_style(cell, StyleFactory.header_style())
        
        running_total = 0
        for i, (name, value) in enumerate(items):
            row = start_row + 2 + i
            ws.cell(row=row, column=start_col, value=name)
            ws.cell(row=row, column=start_col + 1, value=value)
            ws.cell(row=row, column=start_col + 2, value=running_total)
            running_total += value
            ws.cell(row=row, column=start_col + 3, value=running_total)
            ws.cell(row=row, column=start_col + 4, value=value if value > 0 else 0)
            ws.cell(row=row, column=start_col + 5, value=abs(value) if value < 0 else 0)


# =============================================================================
# CONDITIONAL FORMATTING HELPERS
# =============================================================================

class ConditionalFormatHelper:
    """Helpers for applying conditional formatting."""
    
    @staticmethod
    def add_color_scale(
        ws: Worksheet,
        cell_range: str,
        start_color: str = "F8696B",
        mid_color: str = "FFEB84",
        end_color: str = "63BE7B"
    ):
        """Add a 3-color scale conditional format."""
        rule = ColorScaleRule(
            start_type="min", start_color=start_color,
            mid_type="percentile", mid_value=50, mid_color=mid_color,
            end_type="max", end_color=end_color
        )
        ws.conditional_formatting.add(cell_range, rule)
    
    @staticmethod
    def add_data_bars(
        ws: Worksheet,
        cell_range: str,
        color: str = "638EC6"
    ):
        """Add data bars conditional format."""
        rule = DataBarRule(
            start_type="min",
            end_type="max",
            color=color,
            showValue=True,
            minLength=None,
            maxLength=None
        )
        ws.conditional_formatting.add(cell_range, rule)
    
    @staticmethod
    def add_icon_set(
        ws: Worksheet,
        cell_range: str,
        icon_style: str = "3Arrows"
    ):
        """Add icon set conditional format."""
        rule = IconSetRule(
            icon_style=icon_style,
            type="percent",
            values=[0, 33, 67]
        )
        ws.conditional_formatting.add(cell_range, rule)
    
    @staticmethod
    def highlight_cells_greater_than(
        ws: Worksheet,
        cell_range: str,
        value: float,
        fill_color: str = BRAND_COLORS["success_green"]
    ):
        """Highlight cells greater than a value."""
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        rule = CellIsRule(
            operator="greaterThan",
            formula=[str(value)],
            fill=fill
        )
        ws.conditional_formatting.add(cell_range, rule)
    
    @staticmethod
    def highlight_cells_less_than(
        ws: Worksheet,
        cell_range: str,
        value: float,
        fill_color: str = BRAND_COLORS["error_red"]
    ):
        """Highlight cells less than a value."""
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        rule = CellIsRule(
            operator="lessThan",
            formula=[str(value)],
            fill=fill
        )
        ws.conditional_formatting.add(cell_range, rule)


# =============================================================================
# DATA VALIDATION HELPERS
# =============================================================================

class DataValidationHelper:
    """Helpers for creating data validation rules."""
    
    @staticmethod
    def create_dropdown(
        ws: Worksheet,
        cell_range: str,
        options: List[str],
        prompt_title: str = "Select",
        prompt_message: str = "Select from the list"
    ):
        """Create a dropdown list validation."""
        options_str = ",".join(options)
        dv = DataValidation(
            type="list",
            formula1=f'"{options_str}"',
            allow_blank=True
        )
        dv.prompt = prompt_message
        dv.promptTitle = prompt_title
        ws.add_data_validation(dv)
        dv.add(cell_range)
        return dv
    
    @staticmethod
    def create_number_range(
        ws: Worksheet,
        cell_range: str,
        min_val: float,
        max_val: float,
        error_title: str = "Invalid Input",
        error_message: str = "Please enter a valid number"
    ):
        """Create a number range validation."""
        dv = DataValidation(
            type="decimal",
            operator="between",
            formula1=str(min_val),
            formula2=str(max_val),
            allow_blank=True
        )
        dv.error = error_message
        dv.errorTitle = error_title
        ws.add_data_validation(dv)
        dv.add(cell_range)
        return dv
    
    @staticmethod
    def create_percentage_validation(ws: Worksheet, cell_range: str):
        """Create validation for percentage (0-100%)."""
        return DataValidationHelper.create_number_range(
            ws, cell_range, 0, 1,
            "Invalid Percentage",
            "Please enter a value between 0% and 100%"
        )
    
    @staticmethod
    def create_whole_number_validation(
        ws: Worksheet,
        cell_range: str,
        min_val: int = 0,
        max_val: int = 999999999
    ):
        """Create whole number validation."""
        dv = DataValidation(
            type="whole",
            operator="between",
            formula1=str(min_val),
            formula2=str(max_val)
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)
        return dv


# =============================================================================
# WORKSHEET UTILITIES
# =============================================================================

class WorksheetUtils:
    """Utility functions for worksheet operations."""
    
    @staticmethod
    def auto_fit_columns(ws: Worksheet, min_width: int = 8, max_width: int = 50):
        """Auto-fit column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width
    
    @staticmethod
    def set_column_width(ws: Worksheet, column: str, width: float):
        """Set specific column width."""
        ws.column_dimensions[column].width = width
    
    @staticmethod
    def set_row_height(ws: Worksheet, row: int, height: float):
        """Set specific row height."""
        ws.row_dimensions[row].height = height
    
    @staticmethod
    def freeze_panes(ws: Worksheet, cell: str = "B2"):
        """Freeze panes at specified cell."""
        ws.freeze_panes = cell
    
    @staticmethod
    def add_auto_filter(ws: Worksheet, cell_range: str):
        """Add auto filter to a range."""
        ws.auto_filter.ref = cell_range
    
    @staticmethod
    def merge_cells(ws: Worksheet, cell_range: str):
        """Merge cells in a range."""
        ws.merge_cells(cell_range)
    
    @staticmethod
    def unmerge_cells(ws: Worksheet, cell_range: str):
        """Unmerge cells in a range."""
        ws.unmerge_cells(cell_range)
    
    @staticmethod
    def hide_column(ws: Worksheet, column: str):
        """Hide a column."""
        ws.column_dimensions[column].hidden = True
    
    @staticmethod
    def hide_row(ws: Worksheet, row: int):
        """Hide a row."""
        ws.row_dimensions[row].hidden = True
    
    @staticmethod
    def group_columns(ws: Worksheet, start: str, end: str, hidden: bool = True):
        """Group columns with outline."""
        ws.column_dimensions.group(start, end, hidden=hidden)
    
    @staticmethod
    def group_rows(ws: Worksheet, start: int, end: int, hidden: bool = True):
        """Group rows with outline."""
        ws.row_dimensions.group(start, end, hidden=hidden)
    
    @staticmethod
    def add_comment(ws: Worksheet, cell: str, text: str, author: str = ""):
        """Add a comment to a cell."""
        comment = Comment(text, author)
        ws[cell].comment = comment
    
    @staticmethod
    def add_hyperlink(ws: Worksheet, cell: str, url: str, display_text: str = None):
        """Add a hyperlink to a cell."""
        ws[cell].hyperlink = url
        ws[cell].value = display_text or url
        ws[cell].font = Font(color="0563C1", underline="single")
    
    @staticmethod
    def protect_sheet(ws: Worksheet, password: str = None):
        """Protect worksheet."""
        ws.protection.sheet = True
        if password:
            ws.protection.password = password
    
    @staticmethod
    def set_print_area(ws: Worksheet, cell_range: str):
        """Set print area."""
        ws.print_area = cell_range


# =============================================================================
# TABLE BUILDER
# =============================================================================

class TableBuilder:
    """Builder for Excel tables."""
    
    @staticmethod
    def create_table(
        ws: Worksheet,
        table_range: str,
        table_name: str,
        style_name: str = "TableStyleMedium2",
        show_first_column: bool = False,
        show_last_column: bool = False,
        show_row_stripes: bool = True,
        show_column_stripes: bool = False
    ):
        """Create a formatted Excel table."""
        table = Table(displayName=table_name, ref=table_range)
        
        style = TableStyleInfo(
            name=style_name,
            showFirstColumn=show_first_column,
            showLastColumn=show_last_column,
            showRowStripes=show_row_stripes,
            showColumnStripes=show_column_stripes
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        return table
    
    @staticmethod
    def get_available_table_styles() -> List[str]:
        """Get list of available table styles."""
        styles = []
        for color in ["Light", "Medium", "Dark"]:
            max_num = 28 if color != "Dark" else 11
            for i in range(1, max_num + 1):
                styles.append(f"TableStyle{color}{i}")
        return styles


# =============================================================================
# EXCEL SCRIPT EXECUTOR
# =============================================================================

class ExcelScriptExecutor:
    """
    Executes Python scripts that generate Excel workbooks.
    Provides a comprehensive execution environment with all openpyxl features.
    """
    
    def __init__(self):
        self._execution_namespace = self._build_namespace()
    
    def _build_namespace(self) -> Dict[str, Any]:
        """Build the execution namespace with all allowed imports."""
        return {
            # Core openpyxl
            'openpyxl': openpyxl,
            'Workbook': Workbook,
            
            # Styles
            'Font': Font,
            'PatternFill': PatternFill,
            'GradientFill': GradientFill,
            'Alignment': Alignment,
            'Border': Border,
            'Side': Side,
            'Color': Color,
            'Protection': Protection,
            'NamedStyle': NamedStyle,
            
            # Number formats
            'NumberFormat': NumberFormat,
            'NUMBER_FORMATS': NUMBER_FORMATS,
            
            # Utilities
            'get_column_letter': get_column_letter,
            'column_index_from_string': column_index_from_string,
            
            # ALL Chart Types
            'BarChart': BarChart,
            'BarChart3D': BarChart3D,
            'LineChart': LineChart,
            'LineChart3D': LineChart3D,
            'AreaChart': AreaChart,
            'AreaChart3D': AreaChart3D,
            'PieChart': PieChart,
            'PieChart3D': PieChart3D,
            'DoughnutChart': DoughnutChart,
            'ScatterChart': ScatterChart,
            'BubbleChart': BubbleChart,
            'RadarChart': RadarChart,
            'StockChart': StockChart,
            'SurfaceChart': SurfaceChart,
            'SurfaceChart3D': SurfaceChart3D,
            
            # Chart components
            'Reference': Reference,
            'Series': Series,
            'DataLabelList': DataLabelList,
            'Legend': Legend,
            
            # Conditional Formatting
            'ColorScaleRule': ColorScaleRule,
            'FormulaRule': FormulaRule,
            'CellIsRule': CellIsRule,
            'IconSetRule': IconSetRule,
            'DataBarRule': DataBarRule,
            'DifferentialStyle': DifferentialStyle,
            
            # Data validation
            'DataValidation': DataValidation,
            
            # Tables
            'Table': Table,
            'TableStyleInfo': TableStyleInfo,
            
            # Comments
            'Comment': Comment,
            
            # Date/time
            'datetime': datetime,
            'date': date,
            'timedelta': timedelta,
            'relativedelta': relativedelta,
            
            # Helper classes
            'StyleFactory': StyleFactory,
            'ChartFactory': ChartFactory,
            'ChartType': ChartType,
            'FormulaHelper': FormulaHelper,
            'BusinessModelBuilder': BusinessModelBuilder,
            'ConditionalFormatHelper': ConditionalFormatHelper,
            'DataValidationHelper': DataValidationHelper,
            'WorksheetUtils': WorksheetUtils,
            'TableBuilder': TableBuilder,
            
            # Brand colors
            'BRAND_COLORS': BRAND_COLORS,
        }
    
    def validate_script(self, script: str) -> Tuple[bool, str]:
        """Validate script before execution."""
        if not script or not script.strip():
            return False, "Script is empty"
        
        dangerous_patterns = [
            ("import os", "Direct os import is not allowed"),
            ("import sys", "Direct sys import is not allowed"),
            ("import subprocess", "subprocess is not allowed"),
            ("__import__", "Dynamic imports are not allowed"),
            ("eval(", "eval() is not allowed"),
            ("exec(", "Nested exec() is not allowed"),
            ("open(", "File operations are not allowed"),
            ("wb.save(", "wb.save() should not be included - the app handles saving"),
            ("wb.close(", "wb.close() should not be included"),
        ]
        
        for pattern, message in dangerous_patterns:
            if pattern in script:
                return False, message
        
        if "wb" not in script and "Workbook" not in script:
            return False, "Script must create a workbook object named 'wb'"
        
        return True, ""
    
    def execute(self, script: str, validate: bool = True) -> ExecutionResult:
        """Execute a Python script and return the generated workbook."""
        if validate:
            is_valid, error = self.validate_script(script)
            if not is_valid:
                return ExecutionResult(
                    success=False,
                    workbook=None,
                    error_message=error,
                    traceback_str=""
                )
        
        namespace = self._build_namespace().copy()
        
        try:
            exec(script, namespace, namespace)
            
            if 'wb' in namespace:
                workbook = namespace['wb']
                if isinstance(workbook, Workbook):
                    return ExecutionResult(
                        success=True,
                        workbook=workbook,
                        error_message="",
                        traceback_str=""
                    )
                else:
                    return ExecutionResult(
                        success=False,
                        workbook=None,
                        error_message="'wb' is not a valid Workbook object",
                        traceback_str=""
                    )
            else:
                return ExecutionResult(
                    success=False,
                    workbook=None,
                    error_message="Script did not create a workbook object named 'wb'",
                    traceback_str=""
                )
                
        except SyntaxError as e:
            return ExecutionResult(
                success=False,
                workbook=None,
                error_message=f"Syntax error at line {e.lineno}: {e.msg}",
                traceback_str=traceback.format_exc()
            )
        except Exception as e:
            return ExecutionResult(
                success=False,
                workbook=None,
                error_message=str(e),
                traceback_str=traceback.format_exc()
            )
    
    def save_workbook(self, workbook: Workbook, filepath: str) -> Tuple[bool, str]:
        """Save a workbook to file."""
        try:
            workbook.save(filepath)
            return True, ""
        except PermissionError:
            return False, f"Permission denied: Cannot write to {filepath}"
        except Exception as e:
            return False, str(e)
    
    def get_available_features(self) -> Dict[str, List[str]]:
        """Get a summary of available features."""
        return {
            "chart_types": [ct.value for ct in ChartType],
            "number_formats": [nf.name for nf in NumberFormat],
            "helper_classes": [
                "StyleFactory", "ChartFactory", "FormulaHelper",
                "BusinessModelBuilder", "ConditionalFormatHelper",
                "DataValidationHelper", "WorksheetUtils", "TableBuilder"
            ],
            "style_components": [
                "Font", "PatternFill", "GradientFill", "Alignment",
                "Border", "Side", "Color", "Protection", "NamedStyle"
            ],
            "conditional_formatting": [
                "ColorScaleRule", "FormulaRule", "CellIsRule",
                "IconSetRule", "DataBarRule"
            ]
        }


# =============================================================================
# SERVICE INSTANCE
# =============================================================================

excel_executor = ExcelScriptExecutor()
